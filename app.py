import streamlit as st
import streamlit.components.v1 as components
import openpyxl
from supabase import create_client, Client
import json
import time
import base64
import math
from datetime import datetime, date, timedelta

# ==============================================================================
# 🔐 SECURE DATABASE CREDENTIALS BRIDGE
# ==============================================================================
SUPABASE_URL = "https://pysicrdtjayyxztoibep.supabase.co" 
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InB5c2ljcmR0amF5eXh6dG9pYmVwIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODI0Mjk4NzMsImV4cCI6MjA5ODAwNTg3M30.5X0uesuo7NVf6KDxrEiM-6RIOJ2ffyxcOVsWJF52oNw"                    

@st.cache_resource
def get_supabase_client():
    return create_client(SUPABASE_URL, SUPABASE_KEY)

supabase: Client = get_supabase_client()

# ==============================================================================
# 🔌 OFFLINE SYNC TRUCK PROTECTION & AUTOMATED RECONCILIATION ENGINE
# ==============================================================================
def process_bi_directional_sync(local_payload_string, farm_id, current_system_date_str):
    try:
        if not local_payload_string:
            return {"success": False, "msg": "Sync token is empty."}
        payload = json.loads(local_payload_string)
        local_structures = payload.get("structures", [])
        local_logs = payload.get("daily_logs", [])
        
        for item in local_structures:
            for aspect in ["pegging", "piling", "mounting", "modules", "inverter_structure", "inverter", "transformer", "dc_cabling", "ac_cabling"]:
                status_key = f"{aspect}_status"
                date_key = f"{aspect}_date"
                if item.get(status_key) == "completed":
                    if not item.get(date_key) or item.get(date_key) != current_system_date_str:
                        item[date_key] = current_system_date_str
                    supabase.table("structures").update({
                        status_key: "completed", date_key: item[date_key]
                    }).eq("id", item["id"]).execute()
                    
        for log in local_logs:
            supabase.table("daily_progress_logs").upsert({
                "farm_id": str(farm_id), "aspect": log["aspect"], "zone": log["zone"],
                "log_date": log["log_date"], "target_units": int(log["target_units"]),
                "installed_units": int(log["installed_units"]), "deviation": int(log["deviation"]),
                "remark": log.get("remark", "")
            }, on_conflict="farm_id, aspect, zone, log_date").execute()
            
        fresh_structures = supabase.table("structures").select("*").eq("farm_id", farm_id).order("min_r").order("min_c").execute().data
        fresh_schedules = supabase.table("project_schedules").select("*").eq("farm_id", farm_id).execute().data
        fresh_logs = supabase.table("daily_progress_logs").select("*").eq("farm_id", farm_id).order("log_date").execute().data
        
        return {
            "success": True, 
            "bundle": {"structures": fresh_structures or [], "schedules": fresh_schedules or [], "daily_logs": fresh_logs or []}
        }
    except Exception as e:
        return {"success": False, "msg": str(e)}

st.set_page_config(layout="wide", page_title="Boon Solar Farm Tracking System")

# --- HIDE ONLY GITHUB & EDIT PENCIL ICONS ---
st.markdown("""
    <style>
    /* Cleanly collapse and hide the entire upper header toolbar context completely */
    div[data-testid="stAppToolbar"] {
        display: none !important;
    }
    </style>
""", unsafe_allow_html=True)

if "active_site_id" not in st.session_state: st.session_state.active_site_id = None
if "is_admin_mode" not in st.session_state: st.session_state.is_admin_mode = False
if "managed_zones" not in st.session_state: 
    st.session_state.managed_zones = ["Zone A", "Zone B", "Zone C", "Unassigned"]
if "custom_tabs" not in st.session_state: st.session_state.custom_tabs = []

def get_operational_system_date(farm_record):
    if farm_record.get("time_override_date"):
        return datetime.strptime(farm_record["time_override_date"], "%Y-%m-%d").date()
    return date.today()

def calculate_network_working_days(start, end):
    if start > end: return 0
    
    working_days = 0
    current_date = start
    while current_date <= end:
        if current_date.weekday() < 5:
            working_days += 1
        current_date += timedelta(days=1)
        
    return working_days

def fetch_farms_directory():
    try:
        res = supabase.table("farms").select("id, name, admin_password, installer_password, is_published, background_image_url").order("name").execute()
        return res.data if res.data else []
    except Exception: return []

all_registered_farms = fetch_farms_directory()
farm_options = [f["name"] for f in all_registered_farms]

# ==============================================================================
# 🏡 MAIN ENTRY SITE GATEWAY
# ==============================================================================
if st.session_state.active_site_id is None:
    st.title("🚜 Boon Solar Farm Tracking System")
    st.write("---")
    
    with st.sidebar:
        with st.expander("⚙️ Developer Master Control Panel", expanded=False):
            if "dev_unlocked" not in st.session_state:
                st.session_state.dev_unlocked = False

            if not st.session_state.dev_unlocked:
                dev_pwd = st.text_input("Enter Control Password:", type="password")
                if dev_pwd == "devok":
                    st.session_state.dev_unlocked = True
                    st.rerun()
            else:
                st.success("Developer Access Unlocked")
                
                if st.button("🔒 Lock & Exit Developer Mode", type="secondary"):
                    st.session_state.dev_unlocked = False
                    st.rerun()
                
                st.markdown("---")
                st.subheader("🗑️ Cloud Database Cleaner")
                if farm_options:
                    wipe_target = st.selectbox("Select Project to Clear:", farm_options, key="dev_clear_dropdown")
                    
                    if "confirm_purge_gate" not in st.session_state:
                        st.session_state.confirm_purge_gate = False
                    if "purge_target_selected" not in st.session_state:
                        st.session_state.purge_target_selected = ""

                    if st.session_state.purge_target_selected != wipe_target:
                        st.session_state.confirm_purge_gate = False
                        st.session_state.purge_target_selected = wipe_target

                    if not st.session_state.confirm_purge_gate:
                        if st.button("💥 Purge Cloud Data Records", type="primary"):
                            st.session_state.confirm_purge_gate = True
                            st.rerun()
                    else:
                        st.error(f"🚨 **CRITICAL WARNING:** Are you sure you want to delete layout parameters for **{wipe_target}**? This cannot be undone.")
                        
                        col_purge1, col_purge2 = st.columns(2)
                        with col_purge1:
                            if st.button("🔥 YES, PERMANENTLY PURGE", type="primary", use_container_width=True):
                                with st.spinner(f"Purging data assets for {wipe_target}..."):
                                    try:
                                        target_farm = next((f for f in all_registered_farms if f["name"] == wipe_target), None)
                                        if target_farm:
                                            supabase.table("structures").delete().eq("farm_id", target_farm["id"]).execute()
                                            supabase.table("farms").delete().eq("id", target_farm["id"]).execute()
                                            st.success(f"Successfully cleared all data frameworks for {wipe_target}!")
                                            
                                            st.session_state.confirm_purge_gate = False
                                            st.cache_resource.clear()
                                            time.sleep(1)
                                            st.rerun()
                                    except Exception as e: 
                                        st.error(f"Purge rejected: {str(e)}")
                        with col_purge2:
                            if st.button("❌ Cancel / Abort", use_container_width=True):
                                st.session_state.confirm_purge_gate = False
                                st.rerun()
                else:
                    st.info("No active cloud entries found to clear.")
                st.subheader("🚀 Onboard New Layout Framework")
                new_site_name = st.text_input("Assign Site Project Name:")
                init_admin_pwd = st.text_input("Assign Management Password:", value="ok")
                init_inst_pwd = st.text_input("Assign Field Access Password:", value="1234")
                
                uploaded_blueprint = st.file_uploader("Upload Master Blueprint Sheet (.xlsx)", type=["xlsx"])
                
                if uploaded_blueprint and new_site_name and st.button("Compile & Parse Structural Blueprint"):
                    st.info("🔄 Running Fast Visual Grid Scanner...")
                    with st.spinner("Processing structural frames..."):
                        wb = openpyxl.load_workbook(uploaded_blueprint, read_only=False, data_only=True)
                        sheet = wb.active
                        max_rows, max_cols = sheet.max_row, sheet.max_column
                        
                        new_fid = None
                        try:
                            farm_node = supabase.table("farms").insert({
                                "name": new_site_name, "admin_password": init_admin_pwd, "installer_password": init_inst_pwd,
                                "max_rows": max_rows, "max_cols": max_cols, "is_published": False
                            }).execute()
                            if farm_node.data: 
                                new_fid = farm_node.data[0]["id"]
                        except Exception as e:
                            st.error(f"❌ Database registration failed: {str(e)}")
                        
                        if new_fid:
                            grid_matrix = [[False for _ in range(max_cols + 1)] for _ in range(max_rows + 1)]
                            for r in range(1, max_rows + 1):
                                for c in range(1, max_cols + 1):
                                    cell = sheet.cell(row=r, column=c)
                                    if cell.value is not None and str(cell.value).strip() != "":
                                        grid_matrix[r][c] = True
                                    elif cell.fill and cell.fill.fill_type is not None and cell.fill.fill_type != 'none':
                                        grid_matrix[r][c] = True
                                    elif cell.border and ((cell.border.top and cell.border.top.style) or 
                                                         (cell.border.bottom and cell.border.bottom.style) or 
                                                         (cell.border.left and cell.border.left.style) or 
                                                         (cell.border.right and cell.border.right.style)):
                                        grid_matrix[r][c] = True

                            visited_matrix = [[False for _ in range(max_cols + 1)] for _ in range(max_rows + 1)]
                            structures_queue = []
                            table_counter = 1
                            ROAD_GAP = 1

                            for r in range(1, max_rows + 1):
                                for c in range(1, max_cols + 1):
                                    if grid_matrix[r][c] and not visited_matrix[r][c]:
                                        cluster_cells = []
                                        bfs_queue = [(r, c)]
                                        visited_matrix[r][c] = True
                                        discovered_label = ""

                                        while bfs_queue:
                                            curr_r, curr_c = bfs_queue.pop(0)
                                            cluster_cells.append((curr_r, curr_c))

                                            v_cell = sheet.cell(row=curr_r, column=curr_c).value
                                            if v_cell and not discovered_label and not str(v_cell).strip().isdigit():
                                                discovered_label = str(v_cell).strip()

                                            for dr in range(-ROAD_GAP, ROAD_GAP + 1):
                                                for dc in range(-ROAD_GAP, ROAD_GAP + 1):
                                                    nr, nc = curr_r + dr, curr_c + dc
                                                    if 1 <= nr <= max_rows and 1 <= nc <= max_cols:
                                                        if grid_matrix[nr][nc] and not visited_matrix[nr][nc]:
                                                            visited_matrix[nr][nc] = True
                                                            bfs_queue.append((nr, nc))

                                        b_rows = [pt[0] for pt in cluster_cells]
                                        b_cols = [pt[1] for pt in cluster_cells]
                                        min_br, max_br, min_bc, max_bc = min(b_rows), max(b_rows), min(b_cols), max(b_cols)
                                        
                                        h_cells = max_br - min_br + 1
                                        w_cells = max_bc - min_bc + 1

                                        if h_cells >= 2 and w_cells >= 2 and h_cells < 45:
                                            structures_queue.append({
                                                "farm_id": new_fid, 
                                                "table_label": discovered_label if discovered_label else f"T-{table_counter}",
                                                "min_r": int(min_br), "max_r": int(max_br), "min_c": int(min_bc), "max_c": int(max_bc),
                                                "structure_type": "double_6x9" if h_cells >= 5 else "single_3x9",
                                                "assigned_zone": "Unassigned",
                                                "section_group": 403, 
                                                "pegging_status": "pending", "piling_status": "pending", 
                                                "mounting_status": "pending", "modules_status": "pending"
                                            })
                                            table_counter += 1

                            if len(structures_queue) == 0:
                                st.error("❌ Matrix parser rejected configuration: 0 tracker units extracted.")
                            else:
                                success_count = 0
                                for idx in range(0, len(structures_queue), 200):
                                    batch = structures_queue[idx:idx+200]
                                    try: 
                                        supabase.table("structures").insert(batch).execute()
                                        success_count += len(batch)
                                    except Exception: pass
                                
                                st.success(f"🎉 Saved {success_count} structured blocks.")
                                st.cache_resource.clear()
                                time.sleep(1)
                                st.rerun()

    st.subheader("🌐 Access Site Workspace Portal")
    if farm_options:
        with st.form("workspace_access_form"):
            chosen_farm_name = st.selectbox("Select Project Site Layout Location:", farm_options)
            entered_inst_pass = st.text_input("Enter Access Password:", type="password")
            if st.form_submit_button("🚀 Open Workspace"):
                target_site_record = next(f for f in all_registered_farms if f["name"] == chosen_farm_name)
                if str(entered_inst_pass) == str(target_site_record.get("installer_password") or "1234"):
                    st.session_state.active_site_id = target_site_record["id"]
                    st.session_state.active_site_name = target_site_record["name"]
                    st.session_state.admin_key_match = target_site_record.get("admin_password") or "ok"
                    st.session_state.dev_unlocked = False 
                    st.rerun()

# ==============================================================================
# 🗂️ PHASE 2: INTERNAL OPERATIONS TRACKING PLATFORM COMMAND CENTER
# ==============================================================================
else:
    current_farm_record = supabase.table("farms").select("*").eq("id", st.session_state.active_site_id).execute().data[0]
    site_is_published = current_farm_record.get("is_published", False)
    site_bg_img = current_farm_record.get("background_image_url", "")

    current_system_date = get_operational_system_date(current_farm_record)
    current_date_str = str(current_system_date)

    # ==============================================================================
    # 🔐 UNIFIED INTERFACE CONTROL & CREDENTIAL CLEARANCE SIDEBAR BLOCK
    # ==============================================================================
    with st.sidebar:
        st.header("🔐 Workspace Clearances")
        
        if not st.session_state.is_admin_mode:
            with st.form("sidebar_management_credential_verification_form"):
                admin_pass_input = st.text_input(
                    "Enter Management Credentials Panel Pass:", 
                    type="password", 
                    help="Input terminal administrative token keys to authorize project broadcasting layers."
                )
                submit_clearance = st.form_submit_button("Verify Clearance")
                
                if submit_clearance:
                    if admin_pass_input == str(st.session_state.admin_key_match):
                        st.session_state.is_admin_mode = True
                        st.success("Clearance authorized!")
                        time.sleep(0.4)
                        st.rerun()
                    else:
                        st.error("Invalid credentials block.")
                        
        if st.session_state.is_admin_mode:
            st.success("⚡ Admin Permissions Active")
            
            app_theme = st.radio(
                "Visual Dashboard Workspace Theme Profile:", 
                ["Dark Mode", "Light Mode"], 
                key="app_theme_selection_toggle"
            )
            
            st.markdown("---")
            current_system_date = st.date_input(
                "Select System Operation Date Window:", 
                value=current_system_date,
                help="Alters active calendar timeline perspectives for reporting shift segments."
            )
            current_date_str = str(current_system_date)

            st.write("---")
            with st.expander("🧪 Dynamic Workspace Duplicator", expanded=False):
                st.markdown("#### Create an Isolated Testing Sandbox")
                st.caption("Clones your active operational layout matrix completely, including custom stringing and topologies.")
                
                sandbox_suffix = st.text_input("Assign Sandbox Name Extension:", value="EXPERIMENTAL COPY")
                
                if st.button("🚀 Clone Target Setup to New Sandbox Slot", type="primary", use_container_width=True):
                    with st.spinner("Executing 100% Master Duplication Loop..."):
                        try:
                            parent_farm_id = st.session_state.active_site_id
                            raw_topo_string = current_farm_record.get("background_image_url") or "{}"
                            
                            parent_structures = []
                            limit = 1000
                            offset = 0
                            while True:
                                res_page = supabase.table("structures").select("*").eq("farm_id", parent_farm_id).order("min_r").order("min_c").range(offset, offset + limit - 1).execute().data
                                if not res_page: break
                                parent_structures.extend(res_page)
                                if len(res_page) < limit: break
                                offset += limit

                            sandbox_payload = {
                                "name": f"{st.session_state.active_site_name} - {sandbox_suffix.upper()}",
                                "admin_password": current_farm_record.get("admin_password", "ok"),
                                "installer_password": current_farm_record.get("installer_password", "1234"),
                                "max_rows": int(current_farm_record.get("max_rows", 100)),
                                "max_cols": int(current_farm_record.get("max_cols", 150)),
                                "is_published": False,
                                "background_image_url": "{}"
                            }
                            
                            new_farm_response = supabase.table("farms").insert(sandbox_payload).execute()
                            
                            if new_farm_response.data and parent_structures:
                                sandbox_farm_id = new_farm_response.data[0]["id"]
                                
                                sandbox_structures = []
                                for struct in parent_structures:
                                    sandbox_structures.append({
                                        "farm_id": sandbox_farm_id,
                                        "table_label": str(struct.get("table_label", "")),
                                        "min_r": int(struct.get("min_r")), "max_r": int(struct.get("max_r")),
                                        "min_c": int(struct.get("min_c")), "max_c": int(struct.get("max_c")),
                                        "structure_type": str(struct.get("structure_type", "single_3x9")),
                                        "assigned_zone": str(struct.get("assigned_zone", "Unassigned")),
                                        "section_group": int(struct.get("section_group")) if struct.get("section_group") is not None else None,
                                        "pegging_status": str(struct.get("pegging_status", "pending")),
                                        "piling_status": str(struct.get("piling_status", "pending")),
                                        "mounting_status": str(struct.get("mounting_status", "pending")),
                                        "modules_status": str(struct.get("modules_status", "pending"))
                                    })
                                
                                inserted_structures_fleet = []
                                for idx in range(0, len(sandbox_structures), 200):
                                    batch = sandbox_structures[idx:idx+200]
                                    res_batch = supabase.table("structures").insert(batch).execute()
                                    if res_batch.data:
                                        inserted_structures_fleet.extend(res_batch.data)
                                
                                id_mapping_dictionary = {}
                                for old_s in parent_structures:
                                    match = next((new_s for new_s in inserted_structures_fleet 
                                                  if new_s["table_label"] == str(old_s["table_label"]) 
                                                  and int(new_s["min_r"]) == int(old_s["min_r"]) 
                                                  and int(new_s["min_c"]) == int(old_s["min_c"])), None)
                                    if match:
                                        id_mapping_dictionary[str(old_s["id"])] = str(match["id"])
                                
                                try:
                                    if raw_topo_string.startswith("{"):
                                        topo_data = json.loads(raw_topo_string)
                                        
                                        if "stringGroups" in topo_data:
                                            new_string_groups = {}
                                            for old_key, inv_value in topo_data["stringGroups"].items():
                                                parts = old_key.split("_")
                                                old_base_id = parts[0]
                                                suffix = f"_{parts[1]}" if len(parts) > 1 else ""
                                                
                                                if old_base_id in id_mapping_dictionary:
                                                    new_base_id = id_mapping_dictionary[old_base_id]
                                                    new_string_groups[f"{new_base_id}{suffix}"] = inv_value
                                            topo_data["stringGroups"] = new_string_groups
                                        
                                        if "inverters" in topo_data:
                                            CELL = 14
                                            for inv in topo_data["inverters"]:
                                                if "transformerId" in inv and inv["transformerId"] is not None:
                                                    if "transformers" in topo_data and (inv["transformerId"] >= len(topo_data["transformers"]) or inv["transformerId"] < 0):
                                                        inv["transformerId"] = None
                                                
                                                matching_parent_block = None
                                                for p_struct in parent_structures:
                                                    if (p_struct["min_c"] * CELL) <= inv["x"] <= ((p_struct["max_c"] + 1) * CELL) and \
                                                       (p_struct["min_r"] * CELL) <= inv["y"] <= ((p_struct["max_r"] + 1) * CELL):
                                                        matching_parent_block = p_struct
                                                        break
                                                
                                                if matching_parent_block and str(matching_parent_block["id"]) in id_mapping_dictionary:
                                                    new_id_val = int(id_mapping_dictionary[str(matching_parent_block["id"])])
                                                    new_block = next((nb for nb in inserted_structures_fleet if nb["id"] == new_id_val), None)
                                                    if new_block:
                                                        inv["x"] = (new_block["min_c"] * CELL) + (((new_block["max_c"] - new_block["min_c"] + 1) * CELL) / 2)
                                                        inv["y"] = (new_block["min_r"] * CELL) + (((new_block["max_r"] - new_block["min_r"] + 1) * CELL) / 2)
                                                        
                                        raw_topo_string = json.dumps(topo_data)
                                except Exception:
                                    pass
                                
                                supabase.table("farms").update({"background_image_url": raw_topo_string}).eq("id", sandbox_farm_id).execute()
                                
                                st.cache_resource.clear()
                                st.success("🎉 100% Perfect Clone Created! Panels, Inverters, Strings, and TS Hubs are identical.")
                                time.sleep(1.5)
                                st.rerun()
                        except Exception as err:
                            st.error(f"Duplication sequence failed: {str(err)}")
            
            st.write("---")
            st.subheader("📢 Field Deployment Release")
            if not site_is_published:
                st.warning("⚠️ CRITICAL: Review settings carefully. Once published to the field crew, coordinates cannot be altered.")
                
                if "confirm_publish_gate" not in st.session_state: st.session_state.confirm_publish_gate = False
                
                if not st.session_state.confirm_publish_gate:
                    if st.button("🚀 Publish Layout Workspace to Field Crew", type="primary", use_container_width=True):
                        st.session_state.confirm_publish_gate = True
                        st.rerun()
                else:
                    st.error("🔒 Confirm permanent field synchronization lock?")
                    col_lock1, col_lock2 = st.columns(2)
                    with col_lock1:
                        if st.button("🔒 YES, DEPLOY", type="primary", use_container_width=True):
                            supabase.table("farms").update({"is_published": True}).eq("id", st.session_state.active_site_id).execute()
                            st.session_state.confirm_publish_gate = False
                            st.success("Workspace deployed cleanly! Fields locked.")
                            time.sleep(1); st.rerun()
                    with col_lock2:
                        if st.button("Cancel", use_container_width=True):
                            st.session_state.confirm_publish_gate = False
                            st.rerun()
            else:
                st.success("✅ Layout Workspace Status: Locked & Live to Crew")
                if st.button("🔓 Emergency Revoke & Unfreeze Project (Admin Only)", use_container_width=True):
                    supabase.table("farms").update({"is_published": False}).eq("id", st.session_state.active_site_id).execute()
                    st.rerun()
            
            st.write("---")
            st.subheader("🖼️ Map Background Image Configuration")
            uploaded_map_img = st.file_uploader("Upload Layout Image Blueprint (PNG/JPG):", type=["png", "jpg", "jpeg"])
            if uploaded_map_img:
                img_bytes = uploaded_map_img.read()
                b64_img_string = f"data:{uploaded_map_img.type};base64," + base64.b64encode(img_bytes).decode("utf-8")
                
                if site_is_published:
                    st.error("Cannot change image background assets on published, finalized frameworks.")
                elif st.button("💾 Apply & Save Image Blueprint", type="primary", use_container_width=True):
                    supabase.table("farms").update({"background_image_url": b64_img_string}).eq("id", st.session_state.active_site_id).execute()
                    st.success("Image blueprints attached safely!")
                    time.sleep(0.5); st.rerun()
            
            if site_bg_img and not site_is_published:
                if st.button("🗑️ Remove Current Background Image", type="secondary", use_container_width=True):
                    if not site_bg_img.startswith("{"):
                        supabase.table("farms").update({"background_image_url": ""}).eq("id", st.session_state.active_site_id).execute()
                        st.success("Background mapping reference flushed!")
                        time.sleep(0.5); st.rerun()

            st.write("---")
            st.subheader("🛠️ Custom Tracker Tab Builder")
            custom_tab_name = st.text_input("Assign New Tracker Tab Label:", placeholder="e.g. Floating Cell...")
            if st.button("✨ Instantiate Phase Tab", use_container_width=True) and custom_tab_name:
                if custom_tab_name not in st.session_state.custom_tabs:
                    st.session_state.custom_tabs.append(custom_tab_name)
                    st.success(f"Instantiated '{custom_tab_name}'!")
                    time.sleep(0.4); st.rerun()
                    
            if st.button("🔒 Revoke Admin Clearances", use_container_width=True): 
                st.session_state.is_admin_mode = False
                st.rerun()

    # ==============================================================================
    # 🛰️ MAIN DASHBOARD APEX HEADER INTERFACE PANEL 
    # ==============================================================================
    col_h1, col_h2 = st.columns([8, 2])
    with col_h1: 
        st.subheader(f"📍 Boon Solar Farm Tracking System — {st.session_state.active_site_name}")
    with col_h2:
        if st.button("🚪 Exit Site", use_container_width=True): 
            st.session_state.active_site_id = None
            st.session_state.is_admin_mode = False
            st.rerun()
    # ==============================================================================
    # 🔄 LOCAL OFFLINE INTERACTIVE DATA SYNC STATION
    # ==============================================================================
    with st.expander("🔄 Local Workspace Sync Station", expanded=False):
        st.caption("Synchronize your browser offline package data token to pull admin layout updates or upload current shift progress.")
        local_payload_input = st.text_area("Paste Local Browser Sync String Token here:", height=68)
        if st.button("🚀 Reconcile Fleet Operations & Download Updated Plans", type="primary", use_container_width=True):
            if local_payload_input:
                sync_result = process_bi_directional_sync(local_payload_input, st.session_state.active_site_id, current_date_str)
                if sync_result["success"]:
                    st.success("🎉 Bi-directional operational data pipeline successfully resolved!")
                    st.text_area("📋 COPY AND PASTE THIS NEW FRESH TOKEN BACK TO YOUR OFFLINE WORKSPACE COPY:", value=json.dumps(sync_result["bundle"]), height=68)
                else:
                    st.error(f"Sync Interrupted: {sync_result['msg']}")
            
    if not st.session_state.is_admin_mode:
        app_theme = "Dark Mode"

    if st.button("🔄 Reload Workspace Map from Database", type="secondary"):
        st.rerun()

    def load_site_isolated_tables(farm_id):
        all_data = []
        limit = 1000; offset = 0
        while True:
            try:
                res = supabase.table("structures").select("*").eq("farm_id", farm_id).order("min_r").order("min_c").range(offset, offset + limit - 1).execute().data
                if not res: break
                all_data.extend(res)
                if len(res) < limit: break
                offset += limit
            except Exception: break
        return all_data

    active_table_data = load_site_isolated_tables(st.session_state.active_site_id)

    for b in active_table_data:
        z = b.get("assigne_zone")
        if z and z not in st.session_state.managed_zones:
            st.session_state.managed_zones.insert(len(st.session_state.managed_zones)-1, z)
    clean_zones = [z for z in st.session_state.managed_zones if z != "Unassigned"]

    if not active_table_data:
        st.warning("ℹ️ No operational layout metrics have loaded from database for this specific site yet.")
        st.stop()

    min_r = min([b.get("min_r", 1) for b in active_table_data])
    max_r = max([b.get("max_r", 100) for b in active_table_data])
    min_c = min([b.get("min_c", 1) for b in active_table_data])
    max_c = max([b.get("max_c", 150) for b in active_table_data])

    CELL_SIZE = 14
    
    json_str = json.dumps(active_table_data)
    b64_json_data = base64.b64encode(json_str.encode("utf-8")).decode("utf-8")

    for b in active_table_data:
        z = b.get("assigned_zone")
        if z and z not in st.session_state.managed_zones:
            st.session_state.managed_zones.insert(len(st.session_state.managed_zones)-1, z)
    
    clean_wiping_dropdown_options = [zone for zone in st.session_state.managed_zones if zone != "Unassigned"]

    if st.session_state.is_admin_mode:
        setup_tabs = st.tabs([
            "🖼️ Base Overview & Zone Assignation", 
            "📌 Pegging & Piling Customizer",
            "🛰️ Unified Layout Planner & Topology Workspace",
            "📊 Executive Analytical Summary",
            "📅 Schedulers & Targets Manager",  
            "📈 Historical Progress Audit Log"
        ])
        
        # --- STAGE 1: SETUPS OVERVIEW & ZONE ASSIGNATION ---
        with setup_tabs[0]:
            st.markdown("### 🖼️ Operational Field Zoning Assignation Engine")
            
            if site_bg_img and not site_bg_img.startswith("{"):
                st.image(site_bg_img, caption="Active Site Blueprint Layout Background Reference", use_container_width=False, width=600)

            col_z1, col_z2 = st.columns([6, 4])
            with col_z1:
                target_paint_zone = st.selectbox("Active Selector Target Zone Label Options:", st.session_state.managed_zones, index=0)
            with col_z2:
                new_zone_opt = st.text_input("➕ Extend Managed Zone List Registry:", placeholder="e.g. Zone D...")
                if st.button("Register Variant Entry") and new_zone_opt:
                    clean_opt = new_zone_opt.strip()
                    if clean_opt not in st.session_state.managed_zones:
                        st.session_state.managed_zones.insert(len(st.session_state.managed_zones)-1, clean_opt)
                        st.rerun()
            
            st.write("---")
            st.subheader("🛠️ Selective Zone Reset Center")
            col_wipe1, col_wipe2 = st.columns([6, 4])
            with col_wipe1:
                wipe_scope_selection = st.selectbox("Select Target Scope to Flush & Reset to Unassigned:", ["ALL ZONES"] + clean_wiping_dropdown_options)
            with col_wipe2:
                st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
                
                if site_is_published:
                    st.error("Cannot reset zone assets on a frozen, deployed workspace framework.")
                elif st.button("💥 Reset Selected Allocation Fleet", type="secondary", use_container_width=True):
                    with st.spinner("Flushing target zones..."):
                        try:
                            if wipe_scope_selection == "ALL ZONES":
                                supabase.table("structures").update({"assigned_zone": "Unassigned"}).eq("farm_id", st.session_state.active_site_id).execute()
                                st.success("Entire farm structural matrix zones set back to Unassigned!")
                            else:
                                supabase.table("structures").update({"assigned_zone": "Unassigned"}).eq("farm_id", st.session_state.active_site_id).eq("assigned_zone", wipe_scope_selection).execute()
                                st.success(f"Successfully cleared allocation indexes for {wipe_scope_selection}!")
                            time.sleep(0.5); st.rerun()
                        except Exception as e:
                            st.error(f"Reset failed: {str(e)}")

            html_zone_engine = """
            <div style="background:#090d16; padding:12px; border-radius:12px; position:relative; touch-action:none; user-select: none; font-family:sans-serif;">
                <div style="color: #94a3b8; font-size: 13px; margin-bottom: 8px;">
                    Mouse Controls: <span style="color:#22c55e; font-weight:bold;">Left-Click + Drag</span> to select multiple cells &nbsp;|&nbsp; <span style="color:#38bdf8; font-weight:bold;">Right-Click + Drag</span> to pan map &nbsp;|&nbsp; <span style="color:#eab308; font-weight:bold;">Single Left-Click</span> to select a single block &nbsp;|&nbsp; <span style="color:#a78bfa; font-weight:bold;">Scroll</span> to zoom.
                </div>
                
                <div id="canvas_hover_tooltip" style="position: absolute; display: none; background: rgba(15, 23, 42, 0.95); color: #f8fafc; border: 1px solid #38bdf8; padding: 6px 12px; border-radius: 4px; font-size: 12px; pointer-events: none; z-index: 99999; box-shadow: 0 4px 12px rgba(0,0,0,0.5); font-weight: bold;"></div>

                <div id="dialogue_overlay" style="display:none; position:absolute; bottom:35px; left:50%; transform:translateX(-50%); background:#1e293b; padding:18px 35px; border-radius:8px; border:2px solid #38bdf8; z-index:100000; box-shadow: 0 10px 40px rgba(0,0,0,0.85); text-align:center;">
                    <div id="status_message_box" style="color:#22c55e; font-weight:bold; margin-bottom:10px; display:none;">Processing updates...</div>
                    <div style="color:#f1f5f9; font-weight:bold; margin-bottom:14px; font-size:15px;">Assign Selected Section Cluster to <span id="lbl_zone" style="color:#38bdf8; text-decoration:underline;"></span>?</div>
                    <button id="btn_yes" style="background:#22c55e; color:white; border:none; padding:8px 22px; border-radius:4px; font-weight:bold; cursor:pointer; margin-right:12px; font-size:14px;">Yes, Stage Change</button>
                    <button id="btn_no" style="background:#ef4444; color:white; border:none; padding:8px 22px; border-radius:4px; font-weight:bold; cursor:pointer; font-size:14px;">No</button>
                </div>
                <div style="width:100%; max-height:600px; border:2px solid #1e293b; border-radius:8px; overflow:hidden;">
                    <canvas id="zone_canvas" width="1500" height="600" style="background:#020617; display:block;"></canvas>
                </div>
            </div>
            <script>
                (function() {
                    const blocks = JSON.parse(atob("__JSON_DATA_B64__"));
                    const canvas = document.getElementById("zone_canvas");
                    const ctx = canvas.getContext('2d');
                    const tooltip = document.getElementById("canvas_hover_tooltip");
                    const paintZone = "PAINT_ZONE_VAL";
                    const CELL = CELL_SIZE_VAL;
                    const isPublished = __IS_PUBLISHED_VAL__;
                    
                    let minX = MIN_C_VAL, maxX = MAX_C_VAL, minY = MIN_R_VAL, maxY = MAX_R_VAL;
                    const mapWidth = (maxX - minX + 1) * CELL;
                    const mapHeight = (maxY - minY + 1) * CELL;

                    let scale = Math.min((canvas.width - 60) / mapWidth, (canvas.height - 60) / mapHeight);
                    if (scale <= 0 || scale === Infinity) scale = 0.5;

                    let offsetX = (canvas.width / 2) - (mapWidth * scale / 2) - (minX * CELL * scale);
                    let offsetY = (canvas.height / 2) - (mapHeight * scale / 2) - (minY * CELL * scale);

                    let isPanning = false;
                    let isSelecting = false;
                    let startX = 0, startY = 0, currentX = 0, currentY = 0;
                    let stagedBlockIds = [];

                    canvas.addEventListener('contextmenu', e => e.preventDefault());

                    function getZoneColor(zoneName) {
                        if (!zoneName || zoneName.toLowerCase() === 'unassigned' || zoneName.trim() === '') return '#334155';
                        let hash = 0; 
                        let cleaned = zoneName.toUpperCase().trim();
                        for (let i = 0; i < cleaned.length; i++) { 
                            hash = cleaned.charCodeAt(i) + ((hash << 5) - hash); 
                        }
                        let hue = Math.abs(hash * 45) % 360; 
                        return `hsl(${hue}, 90%, 50%)`;
                    }

                    function draw() {
                        ctx.clearRect(0, 0, canvas.width, canvas.height);
                        ctx.save(); ctx.translate(offsetX, offsetY); ctx.scale(scale,scale);

                        blocks.forEach(b => {
                            let isStaged = stagedBlockIds.includes(b.id);
                            ctx.fillStyle = getZoneColor(b.assigned_zone);
                            let x = b.min_c * CELL; let y = b.min_r * CELL;
                            let w = (b.max_c - b.min_c + 1) * CELL; let h = (b.max_r - b.min_r + 1) * CELL;
                            ctx.fillRect(x, y, w, h);
                            ctx.strokeStyle = '#020617'; ctx.lineWidth = 0.75; ctx.strokeRect(x, y, w, h);
                            
                            if (isStaged) { 
                                ctx.strokeStyle = '#ffff00'; 
                                ctx.lineWidth = 2.5; 
                                ctx.strokeRect(x, y, w, h); 
                            }
                        });
                        ctx.restore();

                        if (isSelecting) {
                            ctx.strokeStyle = '#38bdf8'; ctx.lineWidth = 2;
                            ctx.fillStyle = 'rgba(56, 189, 248, 0.25)';
                            ctx.fillRect(startX, startY, currentX - startX, currentY - startY);
                            ctx.strokeRect(startX, startY, currentX - startX, currentY - startY);
                        }
                    }

                    canvas.addEventListener('mousemove', (e) => {
                        const rect = canvas.getBoundingClientRect();
                        const mX = e.clientX - rect.left;
                        const mY = e.clientY - rect.top;
                        
                        if (isPanning) {
                            offsetX = e.clientX - startX;
                            offsetY = e.clientY - startY;
                            draw();
                            tooltip.style.display = "none";
                            return;
                        } else if (isSelecting) {
                            currentX = mX;
                            currentY = mY;
                            draw();
                            tooltip.style.display = "none";
                            return;
                        }

                        let worldX = (mX - offsetX) / scale;
                        let worldY = (mY - offsetY) / scale;
                        
                        let hoveredBlock = null;
                        for (let b of blocks) {
                            let x = b.min_c * CELL; let y = b.min_r * CELL;
                            let w = (b.max_c - b.min_c + 1) * CELL; let h = (b.max_r - b.min_r + 1) * CELL;
                            if (worldX >= x && worldX <= x + w && worldY >= y && worldY <= y + h) {
                                hoveredBlock = b;
                                break;
                            }
                        }

                        if (hoveredBlock) {
                            tooltip.style.display = "block";
                            tooltip.style.left = (mX + 15) + "px";
                            tooltip.style.top = (mY + 15) + "px";
                            tooltip.innerHTML = `Label: ${hoveredBlock.table_label}<br/>Zone: ${hoveredBlock.assigned_zone || 'Unassigned'}`;
                        } else {
                            tooltip.style.display = "none";
                        }
                    });

                    canvas.addEventListener('mousedown', (e) => {
                        if (isPublished) return; 
                        const rect = canvas.getBoundingClientRect();
                        const mX = e.clientX - rect.left;
                        const mY = e.clientY - rect.top;
                        
                        if (e.button === 2) { 
                            isPanning = true;
                            isSelecting = false;
                            startX = e.clientX - offsetX;
                            startY = e.clientY - offsetY;
                            canvas.style.cursor = 'move';
                        } else if (e.button === 0) { 
                            isSelecting = true;
                            isPanning = false;
                            startX = mX;
                            startY = mY;
                            currentX = mX;
                            currentY = mY;
                            canvas.style.cursor = 'crosshair';
                        }
                        tooltip.style.display = "none";
                    });

                    canvas.addEventListener('mouseup', (e) => {
                        const rect = canvas.getBoundingClientRect();
                        const endX = e.clientX - rect.left;
                        const endY = e.clientY - rect.top;

                        if (isPanning) {
                            isPanning = false;
                            canvas.style.cursor = 'grab';
                        } else if (isSelecting) {
                            isSelecting = false;
                            canvas.style.cursor = 'default';
                            
                            stagedBlockIds = [];

                            let boxX1 = Math.min(startX, endX);
                            let boxX2 = Math.max(startX, endX);
                            let boxY1 = Math.min(startY, endY);
                            let boxY2 = Math.max(startY, endY);

                            if (Math.abs(endX - startX) > 4 || Math.abs(endY - startY) > 4) {
                                blocks.forEach(b => {
                                    if (b.assigned_zone && b.assigned_zone.toLowerCase() !== 'unassigned') return;

                                    let cellScreenX1 = b.min_c * CELL * scale + offsetX;
                                    let cellScreenX2 = (b.max_c * CELL + CELL) * scale + offsetX;
                                    let cellScreenY1 = b.min_r * CELL * scale + offsetY;
                                    let cellScreenY2 = (b.max_r * CELL + CELL) * scale + offsetY;

                                    if (cellScreenX2 >= boxX1 && cellScreenX1 <= boxX2 &&
                                        cellScreenY2 >= boxY1 && cellScreenY1 <= boxY2) {
                                        stagedBlockIds.push(b.id);
                                    }
                                });
                            } else {
                                blocks.forEach(b => {
                                    if (b.assigned_zone && b.assigned_zone.toLowerCase() !== 'unassigned') return;

                                    let cellScreenX1 = b.min_c * CELL * scale + offsetX;
                                    let cellScreenX2 = (b.max_c * CELL + CELL) * scale + offsetX;
                                    let cellScreenY1 = b.min_r * CELL * scale + offsetY;
                                    let cellScreenY2 = (b.max_r * CELL + CELL) * scale + offsetY;

                                    if (boxX1 >= cellScreenX1 && boxX1 <= cellScreenX2 &&
                                        boxY1 >= cellScreenY1 && boxY1 <= cellScreenY2) {
                                        stagedBlockIds.push(b.id);
                                    }
                                });
                            }

                            if (stagedBlockIds.length > 0) {
                                document.getElementById("lbl_zone").innerText = paintZone;
                                document.getElementById("dialogue_overlay").style.display = "block";
                            }
                            draw();
                        }
                    });

                    document.getElementById("btn_yes").addEventListener('click', async () => {
                        const msgBox = document.getElementById("status_message_box");
                        msgBox.style.display = "block";
                        msgBox.innerText = `Updating ${stagedBlockIds.length} components...`;
                        
                        try {
                            for (let id of stagedBlockIds) {
                                let target = blocks.find(b => b.id === id); 
                                if (target) target.assigned_zone = paintZone;
                                
                                await fetch("SUPABASE_URL_VAL/rest/v1/structures?id=eq." + id, {
                                    method: "PATCH", 
                                    headers: { 
                                        "apikey": "SUPABASE_KEY_VAL", 
                                        "Authorization": "Bearer SUPABASE_KEY_VAL", 
                                        "Content-Type": "application/json",
                                        "Prefer": "return=minimal"
                                    },
                                    body: JSON.stringify({ "assigned_zone": paintZone })
                                });
                            }
                            msgBox.innerText = "Done! Hit the reload button to refresh overview.";
                            setTimeout(() => { msgBox.style.display = "none"; }, 4000);
                        } catch(e) {
                            msgBox.innerText = "Network transmission exception dropped.";
                        }
                        
                        stagedBlockIds = []; 
                        document.getElementById("dialogue_overlay").style.display = "none"; 
                        draw();
                    });

                    document.getElementById("btn_no").addEventListener('click', () => {
                        stagedBlockIds = []; document.getElementById("dialogue_overlay").style.display = "none"; draw();
                    });

                    canvas.addEventListener('wheel', (e) => {
                        e.preventDefault(); const rect = canvas.getBoundingClientRect(); const mouseX = e.clientX - rect.left; const mouseY = e.clientY - rect.top;
                        const gridX = (mouseX - offsetX) / scale; const gridY = (mouseY - offsetY) / scale;
                        scale *= (e.deltaY < 0 ? 1.15 : 0.85); scale = Math.max(0.005, Math.min(scale, 30));
                        offsetX = mouseX - gridX * scale; offsetY = mouseY - gridY * scale; draw();
                    }, { passive: false });
                    draw();
                })();
            </script>
            """
            html_zone_engine = html_zone_engine.replace("__JSON_DATA_B64__", b64_json_data)\
                                             .replace("PAINT_ZONE_VAL", str(target_paint_zone))\
                                             .replace("CELL_SIZE_VAL", str(CELL_SIZE))\
                                             .replace("MIN_C_VAL", str(min_c))\
                                             .replace("MAX_C_VAL", str(max_c))\
                                             .replace("MIN_R_VAL", str(min_r))\
                                             .replace("MAX_R_VAL", str(max_r))\
                                             .replace("SUPABASE_URL_VAL", SUPABASE_URL)\
                                             .replace("SUPABASE_KEY_VAL", SUPABASE_KEY)\
                                             .replace("__IS_PUBLISHED_VAL__", "true" if site_is_published else "false")
            components.html(html_zone_engine, height=700)

        # --- STAGE 2: PEGGING & PILING CUSTOMIZER ---
        with setup_tabs[1]:
            st.markdown("### 📌 Component Placement Microscale Engineering Template Engine")
            
            layout_types = {}
            for block in active_table_data:
                h_cells = block.get("max_r", 1) - block.get("min_r", 1) + 1
                w_cells = block.get("max_c", 1) - block.get("min_c", 1) + 1
                l_type = block.get("structure_type", "single_3x9")
                layout_key = f"{l_type} ({h_cells}x{w_cells} grid layout)"
                
                if layout_key not in layout_types:
                    layout_types[layout_key] = {
                        "type_string": l_type,
                        "h_cells": h_cells,
                        "w_cells": w_cells,
                        "sample_block_id": block.get("id"),
                        "encoded_value": block.get("section_group") if block.get("section_group") is not None else 403
                    }

            layout_count = len(layout_types)
            if layout_count == 0:
                st.info("No structure architecture frameworks detected.")
            else:
                st.markdown("#### 🗺️ Current Active Database Layout Formations")
                top_cols = st.columns(max(layout_count, 2))
                
                for idx, (label, data) in enumerate(layout_types.items()):
                    with top_cols[idx]:
                        st.markdown(f"📦 **Current saved pattern for {data['type_string'].upper()}**")
                        
                        enc_val = data["encoded_value"]
                        if enc_val > 100:
                            saved_rows = int(enc_val // 100)
                            saved_cols = int(enc_val % 100)
                        else:
                            if enc_val == 12: saved_rows, saved_cols = 3, 4
                            elif enc_val == 6: saved_rows, saved_cols = 2, 3
                            else: saved_rows, saved_cols = 4, 3
                        
                        h_px = int(data["h_cells"] * 18)
                        w_px = int(data["w_cells"] * 26)
                        total_pts = int(saved_rows * saved_cols)

                        html_current_view = f"""
                        <div style="background:#0f172a; padding:10px; border-radius:8px; text-align:center; font-family:sans-serif;">
                            <canvas id="saved_canvas_{idx}" width="320" height="200" style="background:#020617; border:2px solid #22c55e; border-radius:6px;"></canvas>
                            <div style="color:#64748b; font-size:11px; margin-top:4px;">Database Value: {total_pts} Pins ({saved_rows} Rows × {saved_cols} Columns)</div>
                        </div>
                        <script>
                            (function() {{
                                const canvas = document.getElementById("saved_canvas_{idx}");
                                const ctx = canvas.getContext('2d');
                                const rows = {saved_rows};
                                const cols = {saved_cols};
                                const boxW = {w_px}; const boxH = {h_px};
                                
                                const bx = (canvas.width / 2) - (boxW / 2); 
                                const by = (canvas.height / 2) - (boxH / 2);
                                
                                ctx.fillStyle = '#1e293b'; ctx.fillRect(bx, by, boxW, boxH);
                                ctx.strokeStyle = '#22c55e'; ctx.lineWidth = 2; ctx.strokeRect(bx, by, boxW, boxH);
                                
                                if(rows > 0 && cols > 0) {{
                                    const rowGap = (rows === 1) ? boxH / 2 : boxH / (rows - 1);
                                    const colGap = (cols === 1) ? boxW / 2 : boxW / (cols - 1);
                                    
                                    for(let r = 0; r < rows; r++) {{
                                        for(let c = 0; c < cols; c++) {{
                                            let px = (cols === 1) ? bx + (boxW / 2) : bx + (c * colGap);
                                            let py = (rows === 1) ? by + (boxH / 2) : by + (r * rowGap);
                                            
                                            ctx.fillStyle = '#22c55e';
                                            ctx.beginPath();
                                            ctx.arc(px, py, 4.5, 0, Math.PI * 2);
                                            ctx.fill();
                                            ctx.strokeStyle = '#ffffff';
                                            ctx.lineWidth = 1;
                                            ctx.stroke();
                                        }}
                                    }}
                                }}
                            }})();
                        </script>
                        """
                        components.html(html_current_view, height=220)

                st.write("---")
                
                st.markdown("#### ⚙️ Layout Architecture Blueprint Adjustments Deck")
                selected_layout_label = st.selectbox("Select Model Template Variant to Configure Layout Pins Amount:", list(layout_types.keys()))
                
                target_layout = layout_types[selected_layout_label]
                state_prefix = f"layout_cfg_{selected_layout_label}"
                
                if f"{state_prefix}_rows" not in st.session_state: st.session_state[f"{state_prefix}_rows"] = 3
                if f"{state_prefix}_cols" not in st.session_state: st.session_state[f"{state_prefix}_cols"] = 4

                col_inputs, col_actions = st.columns([4, 6])
                
                with col_inputs:
                    row_pts = st.number_input("Array Points per Row Count (Rows stacked vertically):", min_value=1, max_value=20, key=f"{state_prefix}_rows")
                    col_pts = st.number_input("Array Points per Column Count (Columns lined horizontally):", min_value=1, max_value=20, key=f"{state_prefix}_cols")
                    
                    total_calculated_points = row_pts * col_pts
                    st.metric(label="Calculated Configuration Pins", value=f"{total_calculated_points} Pts / Unit")
                    
                    if st.button("💾 Apply & Replicate Fleetwide Structure Patterns", type="primary", use_container_width=True):
                        with st.spinner("Broadcasting layout modifications to cloud ecosystem records..."):
                            try:
                                encoded_group_signature = int((row_pts * 100) + col_pts)
                                supabase.table("structures").update({
                                    "section_group": encoded_group_signature
                                }).eq("farm_id", st.session_state.active_site_id)\
                                  .eq("structure_type", target_layout["type_string"]).execute()
                                
                                st.cache_resource.clear()
                                st.success("Updated fleet configuration profiles cleanly!")
                                time.sleep(0.5)
                                st.rerun()
                            except Exception as e:
                                st.error(f"Transmission mutation failure occurred: {str(e)}")

                with col_actions:
                    h_px_prev = int(target_layout["h_cells"] * 18)
                    w_px_prev = int(target_layout["w_cells"] * 26)
                    
                    html_micro_template = f"""
                    <div style="background:#0f172a; padding:15px; border-radius:12px; text-align:center; font-family:sans-serif;">
                        <div style="margin-bottom: 8px; font-size:12px; color:#94a3b8;">Auto-Updating Local Matrix Preview (Before Saving)</div>
                        <canvas id="micro_canvas" width="400" height="220" style="background:#020617; border:1px dashed #38bdf8; border-radius:8px;"></canvas>
                    </div>
                    <script>
                        (function() {{
                            const canvas = document.getElementById("micro_canvas");
                            const ctx = canvas.getContext('2d');
                            const rows = {row_pts}; const cols = {col_pts};
                            const boxW = {w_px_prev}; const boxH = {h_px_prev};
                            const bx = (canvas.width / 2) - (boxW / 2); const by = (canvas.height / 2) - (boxH / 2);
                            
                            ctx.fillStyle = '#1e293b'; ctx.fillRect(bx, by, boxW, boxH);
                            ctx.strokeStyle = '#38bdf8'; ctx.lineWidth = 2; ctx.strokeRect(bx, by, boxW, boxH);
                            
                            if(rows > 0 && cols > 0) {{
                                const rowGap = (rows === 1) ? boxH / 2 : boxH / (rows - 1);
                                const colGap = (cols === 1) ? boxW / 2 : boxW / (cols - 1);
                                for(let r = 0; r < rows; r++) {{
                                    for(let c = 0; c < cols; c++) {{
                                        let px = (cols === 1) ? bx + (boxW / 2) : bx + (c * colGap);
                                        let py = (rows === 1) ? by + (boxH / 2) : by + (r * rowGap);
                                        ctx.fillStyle = '#f43f5e'; ctx.beginPath(); ctx.arc(px, py, 4.5, 0, Math.PI * 2); ctx.fill();
                                        ctx.strokeStyle = '#ffffff'; ctx.lineWidth = 1; ctx.stroke();
                                    }}
                                }}
                            }}
                        }})();
                    </script>
                    """
                    components.html(html_micro_template, height=270)

        # --- STAGE 3: UNIFIED LAYOUT PLANNER & DC TOPOLOGY WORKSPACE ---
        with setup_tabs[2]:
            st.markdown("### 🔌 Microscale Grid Infrastructure Topologies, Stringing & Drop Station Routing")
            
            stored_metadata_string = current_farm_record.get("background_image_url") if (current_farm_record.get("background_image_url") and current_farm_record.get("background_image_url").startswith("{")) else "{}"
            
            html_topology_workspace = """
            <div style="background:#090d16; padding:15px; border-radius:12px; font-family:sans-serif; color:#f8fafc;">
                <div style="display:grid; grid-template-columns: 260px 1fr; gap:15px;">
                    <div style="background:#0f172a; padding:14px; border-radius:8px; border:1px solid #1e293b; font-size:13px;">
                        <h4 style="margin-top:0; margin-bottom:12px; color:#38bdf8; font-size:14px; border-bottom:1px solid #1e293b; padding-bottom:6px;">🛠️ DESIGN DECK</h4>
                        <label style="display:block; margin-bottom:10px; cursor:pointer;"><input type="radio" name="topo_tool" value="pan" checked> ✋ Pan / Navigate Map</label>
                        <label style="display:block; margin-bottom:10px; cursor:pointer;"><input type="radio" name="topo_tool" value="string"> 🔌 Click / Lasso Strings</label>
                        <label style="display:block; margin-bottom:10px; cursor:pointer;"><input type="radio" name="topo_tool" value="inverter"> ⚡ Click Place Inverter</label>
                        <label style="display:block; margin-bottom:10px; cursor:pointer;"><input type="radio" name="topo_tool" value="transformer"> 🏪 Click Place Transformer</label>
                        <label style="display:block; margin-bottom:10px; cursor:pointer;"><input type="radio" name="topo_tool" value="route"> 🔗 Route Inverters to MVS (Click / Drag Lasso)</label>
                        
                        <hr style="border-color:#1e293b; margin:14px 0;">
                        <h5 style="margin-top:0; margin-bottom:8px; color:#a78bfa; font-size:12px;">ACTIVE IDENTIFICATION</h5>
                        <label style="font-size:11px; color:#94a3b8;">Target Inverter ID #:</label>
                        <input type="number" id="topo_inv_token" value="20" min="1" style="width:100%; background:#1e293b; color:white; border:1px solid #334155; border-radius:4px; padding:5px; margin-bottom:12px; box-sizing:border-box;">
                        
                        <hr style="border-color:#1e293b; margin:14px 0;">
                        <button id="btn_topo_save" style="width:100%; background:#22c55e; border:none; padding:10px 0px; color:white; font-weight:bold; border-radius:4px; cursor:pointer; font-size:13px; line-height:normal; height:auto;">💾 Save Topologies</button>
                    </div>

                    <div style="position:relative;">
                        <div id="topo_tooltip" style="position:absolute; display:none; background:rgba(15,23,42,0.95); border:1px solid #38bdf8; padding:8px; border-radius:4px; font-size:12px; pointer-events:none; z-index:99999; color:#f8fafc; box-shadow:0 4px 12px rgba(0,0,0,0.5);"></div>
                        <canvas id="topo_canvas" width="1120" height="600" style="background:#020617; border-radius:8px; border:1px solid #1e293b; display:block;"></canvas>
                    </div>
                </div>
            </div>

            <script>
                (function() {
                    const databaseStructures = JSON.parse(atob("__JSON_DATA_B64__"));
                    let gridTopo = JSON.parse(atob("__TOPOLOGY_METADATA_B64__"));
                    
                    if (!gridTopo.inverters) gridTopo.inverters = [];
                    if (!gridTopo.transformers) gridTopo.transformers = [];
                    if (!gridTopo.stringGroups) gridTopo.stringGroups = {};

                    const canvas = document.getElementById("topo_canvas");
                    const ctx = canvas.getContext("2d");
                    const tooltip = document.getElementById("topo_tooltip");
                    const CELL = 14;

                    let independentStrings = [];
                    databaseStructures.forEach(b => {
                        if (b.structure_type === "double_6x9") {
                            let mid = b.min_r + Math.floor((b.max_r - b.min_r) / 2);
                            independentStrings.push({
                                id: b.id + "_N", parentId: b.id, label: b.table_label + " (North Facing)",
                                min_c: b.min_c, max_c: b.max_c, min_r: b.min_r, max_r: mid,
                                zone: b.assigned_zone || "Unassigned"
                            });
                            independentStrings.push({
                                id: b.id + "_S", parentId: b.id, label: b.table_label + " (South Facing)",
                                min_c: b.min_c, max_c: b.max_c, min_r: mid + 1, max_r: b.max_r,
                                zone: b.assigned_zone || "Unassigned"
                            });
                        } else {
                            independentStrings.push({
                                id: b.id + "_A", parentId: b.id, label: b.table_label,
                                min_c: b.min_c, max_c: b.max_c, min_r: b.min_r, max_r: b.max_r,
                                zone: b.assigned_zone || "Unassigned"
                            });
                        }
                    });

                    let minX = MIN_C_VAL, maxX = MAX_C_VAL, minY = MIN_R_VAL, maxY = MAX_R_VAL;
                    const mapW = (maxX - minX + 1) * CELL;
                    const mapH = (maxY - minY + 1) * CELL;

                    let scale = Math.min((canvas.width - 80) / mapW, (canvas.height - 80) / mapH);
                    if (scale <= 0 || !isFinite(scale)) scale = 0.5;

                    let offsetX = (canvas.width / 2) - (mapW * scale / 2) - (minX * CELL * scale);
                    let offsetY = (canvas.height / 2) - (mapH * scale / 2) - (minY * CELL * scale);

                    let isPanning = false, isSelecting = false;
                    let startX = 0, startY = 0, currX = 0, currY = 0;
                    let selectedInverterIndexForRouting = null;
                    let lassoSelectedInvertersList = [];

                    canvas.addEventListener("contextmenu", e => e.preventDefault());

                    function getActiveTool() {
                        return document.querySelector('input[name="topo_tool"]:checked').value;
                    }

                    function getCapacityColor(stringCount) {
                        if (stringCount <= 0) return "#1e293b";
                        const palette = [
                            "#10b981", "#06b6d4", "#8b5cf6", "#f43f5e", "#ec4899", 
                            "#3b82f6", "#14b8a6", "#f59e0b", "#6366f1", "#a855f7"
                        ];
                        return palette[(stringCount - 1) % palette.length];
                    }

                    function isInverterPlaced(invId) {
                        return gridTopo.inverters.some(i => i.id === parseInt(invId));
                    }

                    function draw() {
                        ctx.clearRect(0, 0, canvas.width, canvas.height);
                        ctx.save(); ctx.translate(offsetX, offsetY); ctx.scale(scale, scale);

                        gridTopo.inverters.forEach(inv => {
                            if (inv.transformerId !== null && gridTopo.transformers[inv.transformerId]) {
                                let xf = gridTopo.transformers[inv.transformerId];
                                ctx.strokeStyle = "rgba(56, 189, 248, 0.85)"; ctx.lineWidth = 2.0; ctx.beginPath();
                                ctx.moveTo(inv.x, inv.y); ctx.lineTo(xf.x, xf.y); ctx.stroke();
                            }
                        });

                        let counts = {};
                        Object.values(gridTopo.stringGroups).forEach(id => { counts[id] = (counts[id] || 0) + 1; });

                        independentStrings.forEach(s => {
                            let x = s.min_c * CELL; let y = s.min_r * CELL;
                            let w = (s.max_c - s.min_c + 1) * CELL; let h = (s.max_r - s.min_r + 1) * CELL;
                            let linkedInv = gridTopo.stringGroups[s.id];
                            
                            ctx.fillStyle = linkedInv ? (isInverterPlaced(linkedInv) ? getCapacityColor(counts[linkedInv]) : "#d97706") : "#1e293b";
                            ctx.fillRect(x, y, w, h);
                            
                            if (!linkedInv) {
                                ctx.strokeStyle = "rgba(255, 255, 255, 0.12)"; ctx.lineWidth = 0.5; ctx.strokeRect(x, y, w, h);
                            }
                        });

                        let inverterCellsMap = {};
                        independentStrings.forEach(s => {
                            let linkedInv = gridTopo.stringGroups[s.id];
                            if (!linkedInv) return;
                            if (!inverterCellsMap[linkedInv]) inverterCellsMap[linkedInv] = [];
                            inverterCellsMap[linkedInv].push(s);
                        });

                        Object.keys(inverterCellsMap).forEach(invId => {
                            let cellBlocks = inverterCellsMap[invId];
                            ctx.strokeStyle = "#ff0000"; 
                            ctx.lineWidth = 4.0;
                            ctx.lineJoin = "miter";

                            cellBlocks.forEach(b => {
                                let x = b.min_c * CELL;
                                let y = b.min_r * CELL;
                                let w = (b.max_c - b.min_c + 1) * CELL;
                                let h = (b.max_r - b.min_r + 1) * CELL;

                                let topShared = cellBlocks.some(other => b !== other && b.min_r > other.min_r && other.min_c <= b.max_c && other.max_c >= b.min_c && (b.min_r - other.max_r <= 5 || other.parentId === b.parentId));
                                let bottomShared = cellBlocks.some(other => b !== other && b.max_r < other.max_r && other.min_c <= b.max_c && other.max_c >= b.min_c && (other.min_r - b.max_r <= 5 || other.parentId === b.parentId));
                                let leftShared = cellBlocks.some(other => b !== other && b.min_c > other.min_c && other.min_r <= b.max_r && other.max_r >= b.min_r && (b.min_c - other.max_c <= 5));
                                let rightShared = cellBlocks.some(other => b !== other && b.max_c < other.max_c && other.min_r <= b.max_r && other.max_r >= b.min_r && (other.min_c - b.max_c <= 5));

                                ctx.beginPath();
                                if (!topShared) { ctx.moveTo(x, y); ctx.lineTo(x + w, y); }
                                if (!bottomShared) { ctx.moveTo(x, y + h); ctx.lineTo(x + w, y + h); }
                                if (!leftShared) { ctx.moveTo(x, y); ctx.lineTo(x, y + h); }
                                if (!rightShared) { ctx.moveTo(x + w, y); ctx.lineTo(x + w, y + h); }
                                ctx.stroke();
                            });
                        });

                        independentStrings.forEach(s => {
                            let linkedInv = gridTopo.stringGroups[s.id];
                            if (!linkedInv) return;
                            let x = s.min_c * CELL; let y = s.min_r * CELL;
                            ctx.fillStyle = "rgba(0,0,0,0.85)"; ctx.fillRect(x + 2, y + 2, 35, 11);
                            ctx.fillStyle = "#ffffff"; ctx.font = "bold 7px sans-serif";
                            ctx.fillText("I-" + linkedInv, x + 4, y + 10);
                        });

                        gridTopo.inverters.forEach(inv => {
                            let strCount = counts[inv.id] || 0;
                            let tsPrefix = inv.transformerId !== null && gridTopo.transformers[inv.transformerId] ? "TS" + (inv.transformerId + 1) + "-" : "";
                            let titleText = tsPrefix + "IN" + String(inv.id).padStart(3, '0');
                            let countText = strCount + " STRINGS";

                            ctx.font = "bold 9px sans-serif";
                            let badgeW = Math.max(ctx.measureText(titleText).width + 12, ctx.measureText(countText).width + 12, 65);
                            let badgeH = 26; let bx = inv.x - (badgeW / 2); let by = inv.y - (badgeH / 2);

                            ctx.fillStyle = "#ff0000"; ctx.fillRect(bx, by, badgeW, badgeH / 2);
                            ctx.fillStyle = getCapacityColor(strCount); ctx.fillRect(bx, by + (badgeH / 2), badgeW, badgeH / 2);
                            ctx.fillStyle = "#ffffff"; ctx.textAlign = "center";
                            ctx.fillText(titleText, inv.x, by + 9); ctx.fillText(countText, inv.x, by + 21);

                            ctx.strokeStyle = lassoSelectedInvertersList.includes(inv.id) ? "#facc15" : "#ffffff";
                            ctx.lineWidth = lassoSelectedInvertersList.includes(inv.id) ? 3.5 : 1; ctx.strokeRect(bx, by, badgeW, badgeH);
                        });

                        gridTopo.transformers.forEach((t, i) => {
                            ctx.fillStyle = "#ff1744"; ctx.fillRect(t.x - 18, t.y - 18, 36, 36);
                            ctx.strokeStyle = "#ffffff"; ctx.lineWidth = 2; ctx.strokeRect(t.x - 18, t.y - 18, 36, 36);
                            ctx.fillStyle = "#ffffff"; ctx.font = "bold 10px sans-serif"; ctx.textAlign = "center";
                            ctx.fillText("TS " + (i + 1), t.x, t.y + 4);
                        });

                        ctx.restore();

                        if (isSelecting && (getActiveTool() === "string" || getActiveTool() === "route")) {
                            ctx.strokeStyle = getActiveTool() === "string" ? "#a78bfa" : "#38bdf8"; ctx.lineWidth = 1.5;
                            ctx.fillStyle = getActiveTool() === "string" ? "rgba(167, 139, 250, 0.2)" : "rgba(56, 189, 248, 0.2)";
                            ctx.fillRect(startX, startY, currX - startX, currY - startY); ctx.strokeRect(startX, startY, currX - startX, currY - startY);
                        }
                    }

                    function getMouseLocation(e) { const rect = canvas.getBoundingClientRect(); return { x: e.clientX - rect.left, y: e.clientY - rect.top }; }
                    function transformToWorldSpace(p) { return { x: (p.x - offsetX) / scale, y: (p.y - offsetY) / scale }; }

                    canvas.addEventListener("mousedown", e => {
                        const m = getMouseLocation(e); const world = transformToWorldSpace(m); const tool = getActiveTool();

                        if (e.button === 2) {
                            if (tool === "transformer") {
                                let tIdx = gridTopo.transformers.findIndex(t => Math.sqrt(Math.pow(world.x - t.x, 2) + Math.pow(world.y - t.y, 2)) <= 25);
                                if (tIdx !== -1) {
                                    gridTopo.transformers.splice(tIdx, 1);
                                    gridTopo.inverters.forEach(i => { if (i.transformerId === tIdx) i.transformerId = null; else if (i.transformerId > tIdx) i.transformerId -= 1; });
                                }
                                draw();
                                return;
                            }
                            if (tool === "inverter") {
                                gridTopo.inverters = gridTopo.inverters.filter(inv => {
                                    let isHit = Math.sqrt(Math.pow(world.x - inv.x, 2) + Math.pow(world.y - inv.y, 2)) <= 20;
                                    if (isHit) {
                                        Object.keys(gridTopo.stringGroups).forEach(key => {
                                            if (gridTopo.stringGroups[key] === inv.id) delete gridTopo.stringGroups[key];
                                        });
                                    }
                                    return !isHit;
                                });
                                draw(); return;
                            }
                            
                            isPanning = true; startX = e.clientX - offsetX; startY = e.clientY - offsetY; canvas.style.cursor = "move";
                            return;
                        }

                        if (tool === "pan") {
                            isPanning = true; startX = e.clientX - offsetX; startY = e.clientY - offsetY; canvas.style.cursor = "move";
                        } else if (tool === "string" || tool === "route") {
                            if (tool === "route") {
                                let clickedInv = gridTopo.inverters.find(inv => Math.sqrt(Math.pow(world.x - inv.x, 2) + Math.pow(world.y - inv.y, 2)) <= 25);
                                if (clickedInv) {
                                    if (lassoSelectedInvertersList.includes(clickedInv.id)) { lassoSelectedInvertersList = lassoSelectedInvertersList.filter(id => id !== clickedInv.id); }
                                    else { lassoSelectedInvertersList.push(clickedInv.id); }
                                    draw(); return;
                                }
                                let clickedXfmrIdx = gridTopo.transformers.findIndex(t => Math.sqrt(Math.pow(world.x - t.x, 2) + Math.pow(world.y - t.y, 2)) <= 25);
                                if (clickedXfmrIdx !== -1 && lassoSelectedInvertersList.length > 0) {
                                    gridTopo.inverters.forEach(inv => { if (lassoSelectedInvertersList.includes(inv.id)) inv.transformerId = clickedXfmrIdx; });
                                    lassoSelectedInvertersList = []; draw(); return;
                                }
                            }
                            isSelecting = true; startX = m.x; startY = m.y; currX = m.x; currY = m.y;
                        } else if (tool === "inverter") {
                            let hitStruct = databaseStructures.find(b => world.x >= b.min_c*CELL && world.x <= (b.max_c+1)*CELL && world.y >= b.min_r*CELL && world.y <= (b.max_r+1)*CELL);
                            if (hitStruct) {
                                let invId = parseInt(document.getElementById("topo_inv_token").value) || 20;
                                gridTopo.inverters = gridTopo.inverters.filter(i => i.id !== invId);
                                gridTopo.inverters.push({ id: invId, x: (hitStruct.min_c * CELL) + (((hitStruct.max_c - hitStruct.min_c + 1) * CELL) / 2), y: (hitStruct.min_r * CELL) + (((hitStruct.max_r - hitStruct.min_r + 1) * CELL) / 2), transformerId: null });
                                draw();
                            }
                        } else if (tool === "transformer") {
                            let hitStr = independentStrings.find(s => world.x >= s.min_c*CELL && world.x <= (s.max_c+1)*CELL && world.y >= s.min_r*CELL && world.y <= (s.max_r+1)*CELL);
                            if (!hitStr) { gridTopo.transformers.push({ x: world.x, y: world.y }); draw(); }
                        }
                    });

                    canvas.addEventListener("mousemove", e => {
                        const m = getMouseLocation(e); const world = transformToWorldSpace(m);
                        if (isPanning) { offsetX = e.clientX - startX; offsetY = e.clientY - startY; draw(); return; }
                        else if (isSelecting) { currX = m.x; currY = m.y; draw(); return; }

                        let match = false;
                        gridTopo.transformers.forEach((t, idx) => {
                            if (Math.sqrt(Math.pow(world.x - t.x, 2) + Math.pow(world.y - t.y, 2)) <= 25) {
                                let list = gridTopo.inverters.filter(i => i.transformerId === idx).map(i => "INV " + i.id).join(", ");
                                tooltip.style.display = "block"; tooltip.style.left = (m.x+15)+"px"; tooltip.style.top = (m.y+15)+"px";
                                tooltip.innerHTML = `<b>🏪 Transformer Hub</b><br>Station: TS ${idx+1}<br>Fed by: [ ${list || 'None'} ]`; match = true;
                            }
                        });
                        if (!match) {
                            gridTopo.inverters.forEach(inv => {
                                if (Math.sqrt(Math.pow(world.x - inv.x, 2) + Math.pow(world.y - inv.y, 2)) <= 25) {
                                    tooltip.style.display = "block"; tooltip.style.left = (m.x+15)+"px"; tooltip.style.top = (m.y+15)+"px";
                                    tooltip.innerHTML = `<b>⚡ Inverter Node</b><br>ID: ##${inv.id}<br>Route: ${inv.transformerId !== null ? 'TS ' + (inv.transformerId+1) : 'Unassigned'}`; match = true;
                                }
                            });
                        }
                        if (!match) {
                            let s = independentStrings.find(s => world.x >= s.min_c*CELL && world.x <= (s.max_c+1)*CELL && world.y >= s.min_r*CELL && world.y <= (s.max_r+1)*CELL);
                            if (s) {
                                tooltip.style.display = "block"; tooltip.style.left = (m.x+15)+"px"; tooltip.style.top = (m.y+15)+"px";
                                tooltip.innerHTML = `<b>☀️ String Table</b><br>Label: ${s.label}<br>Inverter: ${gridTopo.stringGroups[s.id] || 'None'}`; match = true;
                            }
                        }
                        if (!match) tooltip.style.display = "none";
                    });

                    canvas.addEventListener("mouseup", e => {
                        if (e.button === 2 || isPanning) { isPanning = false; canvas.style.cursor = "default"; return; }
                        if (isSelecting) {
                            isSelecting = false; const mUp = getMouseLocation(e);
                            const p1 = transformToWorldSpace(getMouseLocation({ clientX: startX + canvas.getBoundingClientRect().left, clientY: startY + canvas.getBoundingClientRect().top }));
                            const p2 = transformToWorldSpace(mUp);
                            let x1 = Math.min(p1.x, p2.x), x2 = Math.max(p1.x, p2.x), y1 = Math.min(p1.y, p2.y), y2 = Math.max(p1.y, p2.y);
                            let dist = Math.sqrt(Math.pow(mUp.x - startX, 2) + Math.pow(mUp.y - startY, 2));

                            if (getActiveTool() === "route") {
                                if (dist > 5) {
                                    gridTopo.inverters.forEach(inv => { if (inv.x >= x1 && inv.x <= x2 && inv.y >= y1 && inv.y <= y2) { if (!lassoSelectedInvertersList.includes(inv.id)) lassoSelectedInvertersList.push(inv.id); } });
                                }
                                draw(); return;
                            }

                            let activeInv = parseInt(document.getElementById("topo_inv_token").value) || 20;
                            let isLassoSelection = dist > 5;
                            let boxSelected = independentStrings.filter(s => {
                                let cx = s.min_c * CELL; let cy = s.min_r * CELL;
                                let cw = (s.max_c - s.min_c + 1) * CELL; let ch = (s.max_r - s.min_r + 1) * CELL;
                                if (!isLassoSelection) {
                                    return (x1 >= cx && x1 <= cx + cw && y1 >= cy && y1 <= cy + ch);
                                } else {
                                    return (cx >= x1 && cx <= x2 && cy >= y1 && cy <= y2);
                                }
                            });
                            
                            boxSelected.forEach(s => { if (!gridTopo.stringGroups[s.id] || gridTopo.stringGroups[s.id] === activeInv) { if (!isLassoSelection && gridTopo.stringGroups[s.id] === activeInv) delete gridTopo.stringGroups[s.id]; else gridTopo.stringGroups[s.id] = activeInv; } });
                            draw();
                        }
                    });

                    canvas.addEventListener("wheel", e => {
                        e.preventDefault(); const m = getMouseLocation(e); const world = transformToWorldSpace(m);
                        scale *= (e.deltaY < 0 ? 1.15 : 0.85); scale = Math.max(0.01, Math.min(scale, 20));
                        offsetX = m.x - world.x * scale; offsetY = m.y - world.y * scale; draw();
                    }, { passive: false });

                    draw();
                })();
            </script>
            """
            html_topology_workspace = html_topology_workspace.replace("__JSON_DATA_B64__", b64_json_data)\
                                                             .replace("__TOPOLOGY_METADATA_B64__", base64.b64encode(stored_metadata_string.encode("utf-8")).decode("utf-8"))\
                                                             .replace("MIN_C_VAL", str(min_c))\
                                                             .replace("MAX_C_VAL", str(max_c))\
                                                             .replace("MIN_R_VAL", str(min_r))\
                                                             .replace("MAX_R_VAL", str(max_r))\
                                                             .replace("SUPABASE_URL_VAL", SUPABASE_URL)\
                                                             .replace("SUPABASE_KEY_VAL", SUPABASE_KEY)\
                                                             .replace("ACTIVE_SITE_ID_VAL", str(st.session_state.active_site_id))
            
            components.html(html_topology_workspace, height=660)

        # --- STAGE 4: EXECUTIVE SUMMARY ANALYTICAL MANAGEMENT PANEL TAB ---
        with setup_tabs[3]:
            st.markdown("## 📊 Executive Analytical Summary Dashboard Panel")
            st.write("---")
            
            try: topo_meta = json.loads(current_farm_record.get("background_image_url") or "{}")
            except Exception: topo_meta = {}
            inverters_list = topo_meta.get("inverters", [])
            transformers_list = topo_meta.get("transformers", [])
            string_groups = topo_meta.get("stringGroups", {})
            
            global_inv_string_distribution = {}
            for str_id, inv_id in string_groups.items():
                global_inv_string_distribution[inv_id] = global_inv_string_distribution.get(inv_id, 0) + 1
            
            # Compute Core Structural Dataset Parameters
            layout_analysis = {}
            zone_module_counts = {}
            grand_total_trackers = 0
            grand_total_pegging_points = 0
            grand_total_actual_modules = 0  
            
            for block in active_table_data:
                l_type = block.get("structure_type", "single_3x9")
                z_name = block.get("assigned_zone") if block.get("assigned_zone") else "Unassigned"
                enc_val = block.get("section_group") if block.get("section_group") is not None else 403
                
                if enc_val > 100:
                    r_f = int(enc_val // 100)
                    c_f = int(enc_val % 100)
                    pins_per_unit = int(r_f * c_f)
                else:
                    pins_per_unit = 12; r_f = 4; c_f = 3
                
                # Dynamic panel dimensions tracker math
                grid_rows = int(block["max_r"] - block["min_r"] + 1)
                grid_cols = int(block["max_c"] - block["min_c"] + 1)
                modules_per_tracker = int(grid_rows * grid_cols)
                
                if l_type not in layout_analysis:
                    layout_analysis[l_type] = {
                        "tracker_count": 0, "pins_per_unit": pins_per_unit,
                        "matrix_shape": f"{r_f}×{c_f}", "total_pins": 0, "total_modules": 0
                    }
                    
                layout_analysis[l_type]["tracker_count"] += 1
                layout_analysis[l_type]["total_pins"] += pins_per_unit
                layout_analysis[l_type]["total_modules"] += modules_per_tracker
                
                if z_name not in zone_module_counts:
                    zone_module_counts[z_name] = {"trackers": 0, "pins": 0, "modules": 0}
                zone_module_counts[z_name]["trackers"] += 1
                zone_module_counts[z_name]["pins"] += pins_per_unit
                zone_module_counts[z_name]["modules"] += modules_per_tracker
                
                grand_total_trackers += 1
                grand_total_pegging_points += pins_per_unit
                grand_total_actual_modules += modules_per_tracker

            # ==================================================================
            # LEVEL 1: WHOLE PLANT KEY PERFORMANCE METRICS CARD DECK
            # ==================================================================
            st.subheader("🏭 LEVEL 1: Whole Plant Fleet Summary")
            
            col_p1, col_p2, col_p3, col_p4 = st.columns(4)
            with col_p1: st.metric("⚙️ Total Tracker Tables", f"{grand_total_trackers} Units")
            with col_p2: st.metric("📌 Total Pegging Pins", f"{grand_total_pegging_points} Pts")
            with col_p3: st.metric("📦 Total PV Modules", f"{grand_total_actual_modules} Panels")
            with col_p4: st.metric("⚡ Active Inverter Hubs", f"{len(inverters_list)} INVs")
            
            st.write("")
            
            # Split Layout: Architectural Breakdown Left | Regional Allocation Cards Right
            split_cols = st.columns([5, 5])
            
            with split_cols[0]:
                st.markdown("**Structural Pattern Breakdowns**")
                summary_metrics_rows = []
                for l_name, metrics in layout_analysis.items():
                    summary_metrics_rows.append({
                        "Architecture": l_name.upper(),
                        "Trackers": f"{metrics['tracker_count']} Units",
                        "Matrix": metrics["matrix_shape"],
                        "Pins/Unit": f"{metrics['pins_per_unit']} Pts",
                        "Total Panels": f"{metrics['total_modules']} Modules"
                    })
                st.table(summary_metrics_rows)
                
            with split_cols[1]:
                st.markdown("**Dynamic Modules Count by Zone**")
                sub_cards = st.columns(len(zone_module_counts) if zone_module_counts else 1)
                for idx, z_key in enumerate(sorted(zone_module_counts.keys())):
                    with sub_cards[idx]:
                        st.info(f"**{z_key.upper()}**\n\n🔹 `{zone_module_counts[z_key]['trackers']}` Trackers\n\n🔹 `{zone_module_counts[z_key]['modules']}` Panels")

            # Inline configuration capacity tracker log logic
            global_capacity_buckets = {}
            for inv_id, s_count in global_inv_string_distribution.items():
                bucket_key = f"{s_count} Strings Load"
                if bucket_key not in global_capacity_buckets:
                    global_capacity_buckets[bucket_key] = []
                global_capacity_buckets[bucket_key].append(f"INV #{inv_id}")
                
            if global_capacity_buckets:
                st.write("")
                st.markdown("**Inverter Loading Configuration Breakdown Ledger**")
                g_bucket_rows = []
                for b_name, inv_badge_list in global_capacity_buckets.items():
                    sorted_inv_badges = sorted(inv_badge_list, key=lambda x: int(x.split('#')[1]))
                    g_bucket_rows.append({
                        "String Loading Distribution Capacity": b_name,
                        "Inverters Total Tally": len(sorted_inv_badges),
                        "Mapped Inverter Sequence Pool": ", ".join(sorted_inv_badges)
                    })
                st.table(g_bucket_rows)
            
            st.write("---")
            
            # ==================================================================
            # LEVEL 2: REGIONAL ZONE COMPONENT ALLOCATION METRICS BREAKDOWN
            # ==================================================================
            st.subheader("🗺️ LEVEL 2: Regional Zone Operations Summary Ledger")
            
            # Spatial Matching Engine: Calculate which zone boundary encloses each inverter node
            zone_inverters_map = {z: [] for z in zone_module_counts.keys()}
            CELL = 14
            for inv in inverters_list:
                inv_id = inv.get("id")
                inv_x = inv.get("x", 0)
                inv_y = inv.get("y", 0)
                
                # Identify enclosing spatial coordinate boundaries
                matched_zone = "Unassigned"
                for b in active_table_data:
                    if (b["min_c"] * CELL) <= inv_x <= ((b["max_c"] + 1) * CELL) and \
                       (b["min_r"] * CELL) <= inv_y <= ((b["max_r"] + 1) * CELL):
                        matched_zone = b.get("assigned_zone") if b.get("assigned_zone") else "Unassigned"
                        break
                
                if matched_zone not in zone_inverters_map:
                    zone_inverters_map[matched_zone] = []
                zone_inverters_map[matched_zone].append(inv_id)

            # Build the updated regional configuration ledger table rows
            zone_metrics_rows = []
            for zone_name in sorted(zone_module_counts.keys()):
                metrics = zone_module_counts[zone_name]
                
                # Fetch and format inverters loaded inside this zone boundary
                assigned_inv_ids = sorted(zone_inverters_map.get(zone_name, []))
                inv_count_tally = len(assigned_inv_ids)
                inv_sequence_pool = ", ".join([f"INV #{i}" for i in assigned_inv_ids]) if assigned_inv_ids else "None Routed"
                
                zone_metrics_rows.append({
                    "Zone Sector Area": str(zone_name).upper(),
                    "Total Tracker Tables": f"{metrics['trackers']} Units",
                    "Pegging Pinpoints Total": f"{metrics['pins']} Pts",
                    "Total PV Modules (Panels)": f"{metrics['modules']} Panels",
                    "Active Inverters Count": f"{inv_count_tally} INVs",
                    "Numerical Sequence Inverters Pool": inv_sequence_pool
                })
            st.table(zone_metrics_rows)
            
            st.write("---")
            
            # ==================================================================
            # LEVEL 3: MVS TRANSFORMER STATION INTERCONNECTION REGISTRY
            # ==================================================================
            st.subheader("🏪 LEVEL 3: Transformer Station (MVS) Interconnections")
            st.write(f"**Total Medium Voltage Infrastructure Pool:** `{len(transformers_list)} Active Stations Registered`")
            
            ts_summary_table = []
            for ts_idx, ts_obj in enumerate(transformers_list):
                connected_invs = [inv.get("id") for inv in inverters_list if inv.get("transformerId") == ts_idx]
                connected_invs.sort()
                inv_string_labels = ", ".join([f"INV #{i}" for i in connected_invs]) if connected_invs else "None Routed"
                
                ts_summary_table.append({
                    "Transformer Location ID": f"TS {ts_idx + 1}",
                    "Interconnected Inverters Count": len(connected_invs),
                    "Routed Inverter Hub Sub-Pool": inv_string_labels
                })
            st.table(ts_summary_table)
        
        # --- STAGE 5: SCHEDULERS & TARGETS MANAGER ---
        with setup_tabs[4]:
            st.markdown("### 📅 Aspect & Zone Specific Scheduling Timeline Matrix")
            st.caption("Select an explicit configuration profile below to review or inject custom shift extensions.")
            
            aspect_options_list = ["pegging", "piling", "mounting", "modules", "inverter_structure", "inverter", "transformer", "dc_cabling", "ac_cabling"]
            selected_sched_aspect = st.selectbox("Select Target Aspect Layer Profile:", aspect_options_list, key="admin_sched_aspect_sel")
            
            try: topo_meta_obj = json.loads(current_farm_record.get("background_image_url") or "{}")
            except Exception: topo_meta_obj = {}
            
            zones_to_configure = ["Global"] if selected_sched_aspect == "transformer" else [z for z in st.session_state.managed_zones if z != "Unassigned"]
            selected_target_zone_preview = st.selectbox("Select Target Sector Zone Area:", zones_to_configure, key="zone_preview_filter_idx")
            
            # Master Scope Calculation
            zone_filtered_blocks = [b for b in active_table_data if str(b.get("assigned_zone")) == str(selected_target_zone_preview)]
            zone_specific_element_count = 0
            if selected_sched_aspect in ["pegging", "piling"]:
                for b in zone_filtered_blocks:
                    ev = b.get("section_group") if b.get("section_group") is not None else 403
                    zone_specific_element_count += int((ev // 100) * (ev % 100)) if ev > 100 else 12
            elif selected_sched_aspect in ["mounting", "dc_cabling"]:
                zone_specific_element_count = len(zone_filtered_blocks)
            elif selected_sched_aspect == "modules":
                zone_specific_element_count = sum([int((b["max_r"] - b["min_r"] + 1) * (b["max_c"] - b["min_c"] + 1)) for b in zone_filtered_blocks])
            elif selected_sched_aspect in ["inverter_structure", "inverter", "ac_cabling"]:
                for inv in topo_meta_obj.get("inverters", []):
                    if any(str(b.get("assigned_zone")) == str(selected_target_zone_preview) and (b["min_c"]*14) <= inv.get("x",0) <= ((b["max_c"]+1)*14) and (b["min_r"]*14) <= inv.get("y",0) <= ((b["max_r"]+1)*14) for b in active_table_data):
                        zone_specific_element_count += 1
            elif selected_sched_aspect == "transformer":
                zone_specific_element_count = len(topo_meta_obj.get("transformers", []))

            st.info(f"📊 **Live Spatial Audit:** Found `{zone_specific_element_count}` total items assigned to **{selected_target_zone_preview}**.")

            with st.form("admin_schedule_broadcasting_form"):
                st.markdown("##### 📅 Initialize Base Operations Timeline Boundaries")
                sched_cols = st.columns(2)
                with sched_cols[0]: start_sc_dt = st.date_input("Scheduled Commencement Date:", value=current_system_date)
                with sched_cols[1]: end_sc_dt = st.date_input("Scheduled Operational Wrap Date:", value=current_system_date + timedelta(days=14))
                
                if st.form_submit_button("💾 Broadcast Base Milestones & Save Run-Rates"):
                    w_days = calculate_network_working_days(start_sc_dt, end_sc_dt)
                    if w_days <= 0: 
                        st.error("Terminal exception: End date must exceed start date bounds.")
                    else:
                        computed_runrate_target = round(zone_specific_element_count / w_days, 2)
                        supabase.table("project_schedules").upsert({
                            "farm_id": str(st.session_state.active_site_id), "aspect": str(selected_sched_aspect),
                            "zone": str(selected_target_zone_preview), "start_date": str(start_sc_dt), "end_date": str(end_sc_dt),
                            "working_days": int(w_days), "daily_target": float(computed_runrate_target)
                        }, on_conflict="farm_id, aspect, zone").execute()
                        st.cache_resource.clear()
                        st.success("Milestones initialized successfully!")
                        time.sleep(0.5); st.rerun()

            schedule_meta = supabase.table("project_schedules").select("*").eq("farm_id", st.session_state.active_site_id).eq("aspect", selected_sched_aspect).eq("zone", selected_target_zone_preview).execute().data
            if schedule_meta:
                sched_bound = schedule_meta[0]
                start_bound_dt = datetime.strptime(sched_bound["start_date"], "%Y-%m-%d").date()
                end_bound_dt = datetime.strptime(sched_bound["end_date"], "%Y-%m-%d").date()
                base_working_days = int(sched_bound.get("working_days", 1))
                target_runrate = float(sched_bound.get("daily_target", 0.0))

                raw_logs = supabase.table("daily_progress_logs").select("*").eq("farm_id", st.session_state.active_site_id).eq("aspect", selected_sched_aspect).eq("zone", selected_target_zone_preview).execute().data
                logs_lookup = {r["log_date"]: r for r in raw_logs} if raw_logs else {}
                
                st.write("---")
                st.markdown("#### ⚙️ Manage Weekend Extensions & Special Shifts")
                
                col_add, col_del = st.columns(2)
                with col_add:
                    with st.form("admin_custom_date_injector_form"):
                        st.markdown("##### ➕ Add Shift Extension")
                        extension_date = st.date_input("Select Extension Shift Date:", value=current_system_date)
                        custom_target_quota = st.number_input("Assign FIXED Target Quantity:", min_value=0, value=100)
                        
                        if st.form_submit_button("➕ Inject Shift Day"):
                            if str(extension_date) in logs_lookup and "Special Shift" in (logs_lookup[str(extension_date)].get("remark") or ""):
                                st.error("🛑 Dummy Guard Error: This special shift extension date already exists in the ledger rows!")
                            else:
                                supabase.table("daily_progress_logs").upsert({
                                    "farm_id": str(st.session_state.active_site_id), "aspect": str(selected_sched_aspect),
                                    "zone": str(selected_target_zone_preview), "log_date": str(extension_date),
                                    "target_units": int(custom_target_quota), "installed_units": 0,
                                    "deviation": int(0 - custom_target_quota), "remark": "⚡ Special Shift Authorized by Admin Panel"
                                }, on_conflict="farm_id, aspect, zone, log_date").execute()
                                
                                refreshed_logs = supabase.table("daily_progress_logs").select("*").eq("farm_id", st.session_state.active_site_id).eq("aspect", selected_sched_aspect).eq("zone", selected_target_zone_preview).execute().data
                                total_special_quota = sum(int(l.get("target_units", 0)) for l in refreshed_logs if "Special Shift" in (l.get("remark") or "") or datetime.strptime(l["log_date"], "%Y-%m-%d").date().weekday() >= 5)
                                remaining_base_scope = max(0, int(zone_specific_element_count) - total_special_quota)
                                new_base_daily_target = round(remaining_base_scope / base_working_days, 2)
                                
                                supabase.table("project_schedules").update({"daily_target": float(new_base_daily_target)}).eq("farm_id", str(st.session_state.active_site_id)).eq("aspect", selected_sched_aspect).eq("zone", selected_target_zone_preview).execute()
                                
                                loop_dt = start_bound_dt
                                while loop_dt <= end_bound_dt:
                                    loop_dt_str = str(loop_dt)
                                    if loop_dt.weekday() < 5:
                                        matched = next((r for r in refreshed_logs if r["log_date"] == loop_dt_str), None)
                                        cur_inst = int(matched.get("installed_units") or 0) if matched else 0
                                        cur_rem = matched.get("remark") or "No operational notes submitted." if matched else "No operational notes submitted."
                                        
                                        supabase.table("daily_progress_logs").upsert({
                                            "farm_id": str(st.session_state.active_site_id), "aspect": str(selected_sched_aspect),
                                            "zone": str(selected_target_zone_preview), "log_date": loop_dt_str,
                                            "target_units": int(new_base_daily_target), "installed_units": cur_inst,
                                            "deviation": int(cur_inst - int(new_base_daily_target)), "remark": cur_rem
                                        }, on_conflict="farm_id, aspect, zone, log_date").execute()
                                    loop_dt += timedelta(days=1)
                                    
                                st.cache_resource.clear()
                                st.success("Shift injected & standard day targets redistributed successfully!")
                                time.sleep(0.5); st.rerun()
                                
                with col_del:
                    with st.form("admin_custom_date_removal_form"):
                        st.markdown("##### 🗑️ Remove Shift Extension")
                        special_shift_options = [d for d, l in logs_lookup.items() if "Special Shift" in (l.get("remark") or "") or datetime.strptime(d, "%Y-%m-%d").date().weekday() >= 5]
                        if special_shift_options:
                            target_removal_date = st.selectbox("Select Special Shift to Remove:", special_shift_options)
                            if st.form_submit_button("🗑️ Delete Selected Day Extension", type="primary"):
                                supabase.table("daily_progress_logs").delete().eq("farm_id", st.session_state.active_site_id).eq("aspect", selected_sched_aspect).eq("zone", selected_target_zone_preview).eq("log_date", target_removal_date).execute()
                                
                                refreshed_logs = supabase.table("daily_progress_logs").select("*").eq("farm_id", st.session_state.active_site_id).eq("aspect", selected_sched_aspect).eq("zone", selected_target_zone_preview).execute().data
                                total_special_quota = sum(int(l.get("target_units", 0)) for l in refreshed_logs if "Special Shift" in (l.get("remark") or "") or datetime.strptime(l["log_date"], "%Y-%m-%d").date().weekday() >= 5)
                                remaining_base_scope = max(0, int(zone_specific_element_count) - total_special_quota)
                                new_base_daily_target = round(remaining_base_scope / base_working_days, 2)
                                
                                supabase.table("project_schedules").update({"daily_target": float(new_base_daily_target)}).eq("farm_id", str(st.session_state.active_site_id)).eq("aspect", selected_sched_aspect).eq("zone", selected_target_zone_preview).execute()
                                
                                loop_dt = start_bound_dt
                                while loop_dt <= end_bound_dt:
                                    loop_dt_str = str(loop_dt)
                                    if loop_dt.weekday() < 5:
                                        matched = next((r for r in refreshed_logs if r["log_date"] == loop_dt_str), None)
                                        cur_inst = int(matched.get("installed_units") or 0) if matched else 0
                                        cur_rem = matched.get("remark") or "No operational notes submitted." if matched else "No operational notes submitted."
                                        
                                        supabase.table("daily_progress_logs").upsert({
                                            "farm_id": str(st.session_state.active_site_id), "aspect": str(selected_sched_aspect),
                                            "zone": str(selected_target_zone_preview), "log_date": loop_dt_str,
                                            "target_units": int(new_base_daily_target), "installed_units": cur_inst,
                                            "deviation": int(cur_inst - int(new_base_daily_target)), "remark": cur_rem
                                        }, on_conflict="farm_id, aspect, zone, log_date").execute()
                                    loop_dt += timedelta(days=1)
                                    
                                st.cache_resource.clear()
                                st.success(f"Successfully deleted shift on {target_removal_date} & refreshed baseline targets!")
                                time.sleep(0.5); st.rerun()
                        else:
                            st.info("No removable special shift extensions found in this zone layer.")

                # Mapped Production Operations Calendar Matrix
                st.write("---")
                st.markdown(f"#### 📅 Mapped Production Operations Calendar: `{selected_sched_aspect.upper()}` — `{selected_target_zone_preview.upper()}`")
                
                admin_table_html = """<table style='width:100%; border-collapse: collapse; font-family: sans-serif; text-align: left;'>
                    <thead>
                        <tr style='background-color: #1f2937; color: #f9fafb;'>
                            <th style='padding: 12px; border: 1px solid #374151;'>Date Window</th>
                            <th style='padding: 12px; border: 1px solid #374151;'>Production Target</th>
                            <th style='padding: 12px; border: 1px solid #374151;'>Assembled Quantity</th>
                            <th style='padding: 12px; border: 1px solid #374151;'>Performance Deviation</th>
                            <th style='padding: 12px; border: 1px solid #374151;'>Log Entry Remarks</th>
                        </tr>
                    </thead><tbody>"""

                total_target_quota = 0.0; total_installed_quota = 0.0; total_deviation_quota = 0.0
                loop_date = start_bound_dt
                while loop_date <= end_bound_dt:
                    loop_date_str = str(loop_date)
                    matched_log = logs_lookup.get(loop_date_str)
                    
                    if loop_date.weekday() < 5 or matched_log:
                        inst_count = sum(1 for b in active_table_data if b.get("assigned_zone") == selected_target_zone_preview and b.get(f"{selected_sched_aspect}_status") == "completed" and b.get(f"{selected_sched_aspect}_date") == loop_date_str)
                        t_val = float(matched_log["target_units"]) if (matched_log and matched_log.get("target_units", 0) > 0) else target_runrate
                        cur_dev = float(inst_count - t_val)
                        remark_display = matched_log.get("remark") if matched_log else "No operational notes submitted."
                        
                        if loop_date_str == current_date_str:
                            row_style_attr = "style='background-color: rgba(234, 179, 8, 0.22) !important; font-weight: bold !important; border-left: 5px solid #eab308;'"
                        else:
                            row_style_attr = ""
                        
                        total_target_quota += t_val
                        total_installed_quota += float(inst_count)
                        total_deviation_quota += cur_dev
                        
                        admin_table_html += f"""<tr {row_style_attr}>
                            <td style='padding:10px; border:1px solid #374151;'>{loop_date_str}</td>
                            <td style='padding:10px; border:1px solid #374151;'>{round(t_val)} Units</td>
                            <td style='padding:10px; border:1px solid #374151;'>{int(inst_count)} Units</td>
                            <td style='padding:10px; border:1px solid #374151;'>{"🟢 +" if cur_dev >= 0 else "🔴 "}{round(cur_dev)}</td>
                            <td style='padding:10px; border:1px solid #374151;'>{remark_display}</td>
                        </tr>"""
                    loop_date += timedelta(days=1)
                
                if abs(total_target_quota - zone_specific_element_count) <= 5:
                    total_target_quota = zone_specific_element_count
                    total_deviation_quota = total_installed_quota - total_target_quota

                admin_table_html += f"""<tr style='background:#111827; font-weight:bold;'>
                    <td style='padding:12px; border:1px solid #374151;'>📊 RUNRATES FOOTPRINT ROLLUP TOTALS</td>
                    <td style='padding:12px; border:1px solid #374151;'>{round(total_target_quota)} Units</td>
                    <td style='padding:12px; border:1px solid #374151;'>{round(total_installed_quota)} Units</td>
                    <td style='padding:12px; border:1px solid #374151;'>{"🟢 +" if total_deviation_quota >= 0 else "🔴 "}{round(total_deviation_quota)}</td>
                    <td style='padding:12px; border:1px solid #374151;'>🏁 Master Balance Ledger Log</td>
                </tr></tbody></table>"""
                st.markdown(admin_table_html, unsafe_allow_html=True)

                st.write("---")
                with st.form("admin_schedule_reset_form"):
                    st.markdown("##### 🚨 Danger Zone: Reset Active Schedule")
                    if st.form_submit_button("⚠️ Reset & Clear Active Timeline milestones", type="primary"):
                        supabase.table("project_schedules").delete().eq("farm_id", str(st.session_state.active_site_id)).eq("aspect", selected_sched_aspect).eq("zone", selected_target_zone_preview).execute()
                        supabase.table("daily_progress_logs").delete().eq("farm_id", str(st.session_state.active_site_id)).eq("aspect", selected_sched_aspect).eq("zone", selected_target_zone_preview).execute()
                        st.cache_resource.clear()
                        st.success("Target scheduling vector timeline cleared cleanly!")
                        time.sleep(0.5); st.rerun()

        # --- STAGE 6: HISTORICAL PROGRESS AUDIT LOGS ---
        with setup_tabs[5]:
            st.markdown("### 📈 Historical Operations Playback & Audit Engine Logs")
            
            hist_cols = st.columns(2)
            with hist_cols[0]: lookup_crew_aspect = st.selectbox("Choose Historical Aspect Layer:", ["pegging", "piling", "mounting", "modules", "inverter_structure", "inverter", "transformer", "dc_cabling", "ac_cabling"], key="admin_hist_lookup_asp")
            with hist_cols[1]: lookup_crew_date = st.date_input("Target Evaluation Shift Date Window:", value=current_system_date, key="admin_hist_lookup_dt")
            
            lookup_date_str = str(lookup_crew_date)

            html_hist_zone_map = """
            <div style="background:#090d16; padding:12px; border-radius:12px; position:relative; touch-action:none; user-select: none; font-family:sans-serif;">
                <div style="color: #94a3b8; font-size: 13px; margin-bottom: 8px;">
                    🗺️ <b>Playback Controls:</b> Move cursor to check unit details | <span style="color:#22c55e; font-weight:bold;">Green Cells</span> represent past completions | <span style="color:#eab308; font-weight:bold;">Yellow Cells</span> indicate items installed on this specific shift window. Right-Click + Drag to Pan | Scroll to Zoom.
                </div>
                <div id="admin_hist_canvas_tooltip" style="position: absolute; display: none; background: rgba(15, 23, 42, 0.95); color: #f8fafc; border: 1px solid #3b82f6; padding: 6px 12px; border-radius: 4px; font-size: 12px; pointer-events: none; z-index: 99999; box-shadow: 0 4px 12px rgba(0,0,0,0.5); font-weight: bold;"></div>
                <div style="width:100%; max-height:350px; border:2px solid #1e293b; border-radius:8px; overflow:hidden;">
                    <canvas id="admin_hist_overview_canvas" width="1500" height="350" style="background:#020617; display:block;"></canvas>
                </div>
            </div>
            <script>
                (function() {
                    const blocks = JSON.parse(atob("__JSON_DATA_B64__"));
                    const canvas = document.getElementById("admin_hist_overview_canvas"); const ctx = canvas.getContext('2d'); const CELL = 14;
                    const tooltip = document.getElementById("admin_hist_canvas_tooltip");
                    const aspect = "ACTIVE_ASPECT_VAL"; const evaluationDateStr = "EVALUATION_DATE_VAL";
                    
                    let minX = MIN_C_VAL, maxX = MAX_C_VAL, minY = MIN_R_VAL, maxY = MAX_R_VAL;
                    const mapWidth = (maxX - minX + 1) * CELL; const mapHeight = (maxY - minY + 1) * CELL;
                    let scale = Math.min((canvas.width - 60) / mapWidth, (canvas.height - 60) / mapHeight);
                    let offsetX = (canvas.width / 2) - (mapWidth * scale / 2) - (minX * CELL * scale);
                    let offsetY = (canvas.height / 2) - (mapHeight * scale / 2) - (minY * CELL * scale);
                    let isPanning = false; let startX = 0, startY = 0;
                    
                    function draw() {
                        ctx.clearRect(0, 0, canvas.width, canvas.height); ctx.save(); ctx.translate(offsetX, offsetY); ctx.scale(scale,scale);
                        blocks.forEach(b => {
                            let status = b[aspect + '_status'] || 'pending'; let compDate = b[aspect + '_date'];
                            if (status === 'completed' && compDate) {
                                if (compDate === evaluationDateStr) ctx.fillStyle = '#eab308';
                                else if (compDate < evaluationDateStr) ctx.fillStyle = '#22c55e';
                                else ctx.fillStyle = '#1e293b';
                            } else ctx.fillStyle = '#1e293b';
                            
                            let x = b.min_c * CELL; let y = b.min_r * CELL; let w = (b.max_c - b.min_c + 1) * CELL; let h = (b.max_r - b.min_r + 1) * CELL;
                            ctx.fillRect(x, y, w, h); ctx.strokeStyle = '#020617'; ctx.lineWidth = 0.5; ctx.strokeRect(x, y, w, h);
                        });
                        ctx.restore();
                    }
                    canvas.addEventListener('mousemove', e => {
                        const rect = canvas.getBoundingClientRect(); const mX = e.clientX - rect.left; const mY = e.clientY - rect.top;
                        if (isPanning) { offsetX = e.clientX - startX; offsetY = e.clientY - startY; draw(); tooltip.style.display = "none"; return; }
                        let worldX = (mX - offsetX) / scale; let worldY = (mY - offsetY) / scale;
                        let hoveredBlock = null;
                        for (let b of blocks) {
                            let x = b.min_c * CELL; let y = b.min_r * CELL; let w = (b.max_c - b.min_c + 1) * CELL; let h = (b.max_r - b.min_r + 1) * CELL;
                            if (worldX >= x && worldX <= x + w && worldY >= y && worldY <= y + h) { hoveredBlock = b; break; }
                        }
                        if (hoveredBlock) {
                            tooltip.style.display = "block"; tooltip.style.left = (mX + 15) + "px"; tooltip.style.top = (mY + 15) + "px";
                            let status = hoveredBlock[aspect + '_status'] || 'pending'; let compDate = hoveredBlock[aspect + '_date'] || 'N/A';
                            tooltip.innerHTML = `📍 Unit: ${hoveredBlock.table_label}<br>🗺️ Zone: ${hoveredBlock.assigned_zone}<br>⚡ Status: ${status.toUpperCase()}<br>📅 Built Date: ${compDate}`;
                        } else tooltip.style.display = "none";
                    });
                    canvas.addEventListener('mousedown', e => { if (e.button === 2) { isPanning = true; startX = e.clientX - offsetX; startY = e.clientY - offsetY; tooltip.style.display = "none"; } });
                    canvas.addEventListener('mouseup', () => { isPanning = false; });
                    canvas.addEventListener('wheel', e => {
                        e.preventDefault(); const rect = canvas.getBoundingClientRect(); const mouseX = e.clientX - rect.left; const mouseY = e.clientY - rect.top;
                        const gridX = (mouseX - offsetX) / scale; const gridY = (mouseY - offsetY) / scale;
                        scale *= (e.deltaY < 0 ? 1.15 : 0.85); scale = Math.max(0.01, Math.min(scale, 20));
                        offsetX = mouseX - gridX * scale; offsetY = mouseY - gridY * scale; draw();
                    }, { passive: false });
                    draw();
                })();
            </script>
            """
            html_hist_zone_map = html_hist_zone_map.replace("__JSON_DATA_B64__", b64_json_data).replace("ACTIVE_ASPECT_VAL", lookup_crew_aspect).replace("EVALUATION_DATE_VAL", lookup_date_str).replace("MIN_C_VAL", str(min_c)).replace("MAX_C_VAL", str(max_c)).replace("MIN_R_VAL", str(min_r)).replace("MAX_R_VAL", str(max_r))
            components.html(html_hist_zone_map, height=390)

            st.markdown("#### 📅 Comprehensive Historical Performance Calendars")
            hist_sched_records = supabase.table("project_schedules").select("*").eq("farm_id", st.session_state.active_site_id).eq("aspect", lookup_crew_aspect).execute().data
            
            if hist_sched_records:
                raw_hist_logs = supabase.table("daily_progress_logs").select("*").eq("farm_id", st.session_state.active_site_id).eq("aspect", lookup_crew_aspect).execute().data
                hist_logs_lookup = {(r["zone"], r["log_date"]): r for r in raw_hist_logs} if raw_hist_logs else {}
                
                for zone_name in [z for z in st.session_state.managed_zones if z != "Unassigned"]:
                    sched = next((s for s in hist_sched_records if s["zone"] == zone_name), None)
                    if not sched: continue
                    
                    s_bound = datetime.strptime(sched["start_date"], "%Y-%m-%d").date()
                    e_bound = datetime.strptime(sched["end_date"], "%Y-%m-%d").date()
                    t_runrate = float(sched.get("daily_target", 0.0))
                    
                    st.markdown(f"##### 🗺️ Performance Metrics Log Schedule: **{zone_name.upper()}**")
                    
                    hist_table_html = """<table style='width:100%; border-collapse: collapse; font-family: sans-serif; text-align: left; margin-bottom: 24px;'>
                        <thead>
                            <tr style='background-color: #1f2937; color: #f9fafb;'>
                                <th style='padding: 12px; border: 1px solid #374151;'>Date Window</th>
                                <th style='padding: 12px; border: 1px solid #374151;'>Production Target</th>
                                <th style='padding: 12px; border: 1px solid #374151;'>Actual Installed</th>
                                <th style='padding: 12px; border: 1px solid #374151;'>Shift Performance Deviation</th>
                                <th style='padding: 12px; border: 1px solid #374151;'>Supervisor Operational Remarks</th>
                            </tr>
                        </thead><tbody>"""

                    z_total_target = 0.0; z_total_installed = 0.0; z_total_deviation = 0.0
                    
                    # Collate specific date index arrays cleanly to safely kill any NameErrors
                    all_zone_dates = set([l_key[1] for l_key in hist_logs_lookup.keys() if l_key[0] == zone_name])
                    c_cursor = s_bound
                    while c_cursor <= e_bound:
                        if c_cursor.weekday() < 5:
                            all_zone_dates.add(str(c_cursor))
                        c_cursor += timedelta(days=1)
                        
                    for loop_dt_str in sorted(list(all_zone_dates)):
                        matched_log = hist_logs_lookup.get((zone_name, loop_dt_str))
                        inst_val = sum(1 for b in active_table_data if b.get("assigned_zone") == zone_name and b.get(f"{lookup_crew_aspect}_status") == "completed" and b.get(f"{lookup_crew_aspect}_date") == loop_dt_str)
                        
                        t_val = float(matched_log["target_units"]) if (matched_log and matched_log.get("target_units", 0) > 0) else t_runrate
                        dev_val = float(inst_val - t_val)
                        rem_val = matched_log.get("remark") or "No field remarks submitted." if matched_log else "No logs recorded."
                        
                        if loop_dt_str == lookup_date_str:
                            row_style = "style='background-color: rgba(59, 130, 246, 0.24) !important; font-weight: bold !important; border-left: 5px solid #3b82f6 !important;'"
                        else:
                            row_style = ""
                            
                        z_total_target += t_val
                        z_total_installed += float(inst_val)
                        z_total_deviation += dev_val
                        
                        hist_table_html += f"""<tr {row_style}>
                            <td style='padding: 10px; border: 1px solid #374151;'>{loop_dt_str}</td>
                            <td style='padding: 10px; border: 1px solid #374151;'>{round(t_val)} Units</td>
                            <td style='padding: 10px; border: 1px solid #374151;'>{int(inst_val)} Units</td>
                            <td style='padding: 10px; border: 1px solid #374151;'>{"🟢 +" if dev_val >= 0 else "🔴 "}{round(dev_val)}</td>
                            <td style='padding: 10px; border: 1px solid #374151;'>{rem_val}</td>
                        </tr>"""
                        
                    hist_table_html += f"""<tr style='background:#111827; font-weight:bold;'>
                        <td style='padding:12px; border: 1px solid #374151;'>📊 {zone_name.upper()} SUMMATION ROLLUP TOTALS</td>
                        <td style='padding:12px; border: 1px solid #374151;'>{round(z_total_target)} Units</td>
                        <td style='padding:12px; border: 1px solid #374151;'>{round(z_total_installed)} Units</td>
                        <td style='padding:12px; border: 1px solid #374151;'>{"🟢 +" if z_total_deviation >= 0 else "🔴 "}{round(z_total_deviation)}</td>
                        <td style='padding:12px; border: 1px solid #374151;'>🏁 Zone Specific Ledger Balance</td>
                    </tr></tbody></table>"""
                    st.markdown(hist_table_html, unsafe_allow_html=True)
            else:
                st.info("ℹ️ No active milestone configuration layers match this aspect layer context.")

    # ==============================================================================
    # 👷 LIVE PRODUCTION CREW WORKSPACE MAPPING DASHBOARDS
    # ==============================================================================
    else:
        st.markdown("### 🛰️ Live Production Crew Workspace Mapping Dashboards")
        
        crew_tabs = st.tabs([
            "🗺️ Whole Plant Master Blueprint Index", 
            "🛠️ Execution Workspace Tracker Deck",
            "🕒 Field Shift History Log Viewer"
        ])
        
        # --- CREW TAB 1: WHOLE FARM OVERVIEW TAB ---
        with crew_tabs[0]:
            st.markdown("### 🖼️ Whole Plant Spatial Allocation & Fleet Metrics Index")
            
            html_crew_zone_map = """
            <div style="background:#090d16; padding:12px; border-radius:12px; position:relative; touch-action:none; user-select: none; font-family:sans-serif;">
                <div style="color: #94a3b8; font-size: 13px; margin-bottom: 8px;">
                    ✋ <b>Map Controls:</b> Move your cursor across cells to see Zone details | Right-Click + Drag to Pan &nbsp;|&nbsp; Scroll to Zoom.
                </div>
                <div id="crew_overview_tooltip" style="position: absolute; display: none; background: rgba(15, 23, 42, 0.95); color: #f8fafc; border: 1px solid #38bdf8; padding: 6px 12px; border-radius: 4px; font-size: 12px; pointer-events: none; z-index: 99999; box-shadow: 0 4px 12px rgba(0,0,0,0.5); font-weight: bold;"></div>
                <div style="width:100%; max-height:400px; border:2px solid #1e293b; border-radius:8px; overflow:hidden;">
                    <canvas id="crew_zone_overview_canvas" width="1500" height="400" style="background:#020617; display:block;"></canvas>
                </div>
            </div>
            <script>
                (function() {
                    const blocks = JSON.parse(atob("__JSON_DATA_B64__"));
                    const canvas = document.getElementById("crew_zone_overview_canvas"); const ctx = canvas.getContext('2d'); const CELL = 14;
                    const tooltip = document.getElementById("crew_overview_tooltip");
                    let minX = MIN_C_VAL, maxX = MAX_C_VAL, minY = MIN_R_VAL, maxY = MAX_R_VAL;
                    const mapWidth = (maxX - minX + 1) * CELL; const mapHeight = (maxY - minY + 1) * CELL;
                    let scale = Math.min((canvas.width - 60) / mapWidth, (canvas.height - 60) / mapHeight);
                    let offsetX = (canvas.width / 2) - (mapWidth * scale / 2) - (minX * CELL * scale);
                    let offsetY = (canvas.height / 2) - (mapHeight * scale / 2) - (minY * CELL * scale);
                    let isPanning = false; let startX = 0, startY = 0;
                    
                    function getZoneColor(zoneName) {
                        if (!zoneName || zoneName.toLowerCase() === 'unassigned') return '#1f2937';
                        let hash = 0; let cleaned = zoneName.toUpperCase().trim();
                        for (let i = 0; i < cleaned.length; i++) { hash = cleaned.charCodeAt(i) + ((hash << 5) - hash); }
                        return `hsl(${Math.abs(hash * 45) % 360}, 75%, 45%)`;
                    }
                    
                    function draw() {
                        ctx.clearRect(0, 0, canvas.width, canvas.height); ctx.save(); ctx.translate(offsetX, offsetY); ctx.scale(scale,scale);
                        let zoneCenters = {};
                        blocks.forEach(b => {
                            ctx.fillStyle = getZoneColor(b.assigned_zone);
                            let x = b.min_c * CELL; let y = b.min_r * CELL; let w = (b.max_c - b.min_c + 1) * CELL; let h = (b.max_r - b.min_r + 1) * CELL;
                            ctx.fillRect(x, y, w, h); ctx.strokeStyle = '#020617'; ctx.lineWidth = 0.5; ctx.strokeRect(x, y, w, h);
                            if (b.assigned_zone && b.assigned_zone.toLowerCase() !== 'unassigned') {
                                if (!zoneCenters[b.assigned_zone]) zoneCenters[b.assigned_zone] = { sumX: 0, sumY: 0, count: 0 };
                                zoneCenters[b.assigned_zone].sumX += x + (w / 2); zoneCenters[b.assigned_zone].sumY += y + (h / 2); zoneCenters[b.assigned_zone].count++;
                            }
                        });
                        Object.keys(zoneCenters).forEach(zName => {
                            let c = zoneCenters[zName]; let centerX = c.sumX / c.count; let centerY = c.sumY / c.count;
                            ctx.fillStyle = "rgba(15, 23, 42, 0.85)"; ctx.font = "bold 13px sans-serif";
                            let textWidth = ctx.measureText(zName.toUpperCase()).width;
                            ctx.fillRect(centerX - (textWidth/2) - 8, centerY - 10, textWidth + 16, 20);
                            ctx.fillStyle = "#ffffff"; ctx.textAlign = "center"; ctx.fillText(zName.toUpperCase(), centerX, centerY + 4);
                        });
                        ctx.restore();
                    }
                    canvas.addEventListener('mousemove', e => {
                        const rect = canvas.getBoundingClientRect(); const mX = e.clientX - rect.left; const mY = e.clientY - rect.top;
                        if (isPanning) { offsetX = e.clientX - startX; offsetY = e.clientY - startY; draw(); tooltip.style.display = "none"; return; }
                        let worldX = (mX - offsetX) / scale; let worldY = (mY - offsetY) / scale;
                        let hoveredBlock = null;
                        for (let b of blocks) {
                            let x = b.min_c * CELL; let y = b.min_r * CELL; let w = (b.max_c - b.min_c + 1) * CELL; let h = (b.max_r - b.min_r + 1) * CELL;
                            if (worldX >= x && worldX <= x + w && worldY >= y && worldY <= y + h) { hoveredBlock = b; break; }
                        }
                        if (hoveredBlock) {
                            tooltip.style.display = "block"; tooltip.style.left = (mX + 15) + "px"; tooltip.style.top = (mY + 15) + "px";
                            tooltip.innerHTML = `📍 Label: ${hoveredBlock.table_label}<br/>🗺️ Zone: ${hoveredBlock.assigned_zone || 'Unassigned'}`;
                        } else tooltip.style.display = "none";
                    });
                    canvas.addEventListener('mousedown', e => { if (e.button === 2) { isPanning = true; startX = e.clientX - offsetX; startY = e.clientY - offsetY; tooltip.style.display = "none"; } });
                    canvas.addEventListener('mouseup', () => { isPanning = false; });
                    canvas.addEventListener('wheel', e => {
                        e.preventDefault(); const rect = canvas.getBoundingClientRect(); const mouseX = e.clientX - rect.left; const mouseY = e.clientY - rect.top;
                        const gridX = (mouseX - offsetX) / scale; const gridY = (mouseY - offsetY) / scale;
                        scale *= (e.deltaY < 0 ? 1.15 : 0.85); scale = Math.max(0.01, Math.min(scale, 20));
                        offsetX = mouseX - gridX * scale; offsetY = mouseY - gridY * scale; draw();
                    }, { passive: false });
                    draw();
                })();
            </script>
            """
            html_crew_zone_map = html_crew_zone_map.replace("__JSON_DATA_B64__", b64_json_data).replace("MIN_C_VAL", str(min_c)).replace("MAX_C_VAL", str(max_c)).replace("MIN_R_VAL", str(min_r)).replace("MAX_R_VAL", str(max_r))
            components.html(html_crew_zone_map, height=440)

            # Executive Metrics summary Cards Row
            try: topo_meta = json.loads(current_farm_record.get("background_image_url") or "{}")
            except Exception: topo_meta = {}
            inverters_list = topo_meta.get("inverters", [])
            st.markdown("#### 🏭 Executive Regional Allocation Summary Ledger")
            col_m1, col_p2, col_p3, col_p4 = st.columns(4)
            grand_total_trackers = len(active_table_data)
            grand_total_pins = sum([int((b["section_group"] // 100) * (b["section_group"] % 100)) if (b.get("section_group") and b["section_group"] > 100) else 12 for b in active_table_data])
            grand_total_panels = sum([int((b["max_r"] - b["min_r"] + 1) * (b["max_c"] - b["min_c"] + 1)) for b in active_table_data])
            with col_m1: st.metric("⚙️ Total Tracker Tables", f"{grand_total_trackers} Units")
            with col_p2: st.metric("📌 Total Pegging Pins", f"{grand_total_pins} Pts")
            with col_p3: st.metric("📦 Total PV Modules", f"{grand_total_panels} Panels")
            with col_p4: st.metric("⚡ Active Inverter Hubs", f"{len(inverters_list)} INVs")
            
            zone_analysis = {}
            zone_inverters_map = {z: [] for z in clean_zones}
            for block in active_table_data:
                z_name = block.get("assigned_zone") if block.get("assigned_zone") else "Unassigned"
                enc_val = block.get("section_group") if block.get("section_group") is not None else 403
                pins = int(enc_val // 100 * (enc_val % 100)) if enc_val > 100 else 12
                panels = int((block["max_r"] - block["min_r"] + 1) * (block["max_c"] - block["min_c"] + 1))
                if z_name not in zone_analysis: zone_analysis[z_name] = {"trackers": 0, "pins": 0, "modules": 0}
                zone_analysis[z_name]["trackers"] += 1; zone_analysis[z_name]["pins"] += pins; zone_analysis[z_name]["modules"] += panels

            for inv in inverters_list:
                inv_x, inv_y = inv.get("x", 0), inv.get("y", 0)
                matched_z = "Unassigned"
                for b in active_table_data:
                    if (b["min_c"] * 14) <= inv_x <= ((b["max_c"] + 1) * 14) and (b["min_r"] * 14) <= inv_y <= ((b["max_r"] + 1) * 14):
                        matched_z = b.get("assigned_zone") or "Unassigned"; break
                if matched_z in zone_inverters_map: zone_inverters_map[matched_z].append(inv.get("id"))

            executive_rows = []
            for zone_name, metrics in sorted(zone_analysis.items()):
                if zone_name == "Unassigned": continue
                invs_pool = zone_inverters_map.get(zone_name, [])
                executive_rows.append({
                    "Zone Sector Area Context": zone_name.upper(), "Total Tracker Tables": f"{metrics['trackers']} Units",
                    "Pegging Pinpoints Total": f"{metrics['pins']} Pts", "Total PV Modules (Panels)": f"{metrics['modules']} Panels",
                    "Active Inverters Pool": f"{len(invs_pool)} INVs (" + (", ".join([f"INV #{i}" for i in sorted(invs_pool)]) if invs_pool else "None") + ")"
                })
            st.table(executive_rows)

            # Master schedule grid block
            st.markdown("#### 📅 Whole Plant Master Schedule Matrix Index")
            saved_schedules_res = supabase.table("project_schedules").select("*").eq("farm_id", st.session_state.active_site_id).execute().data
            sched_lookup = {(r["aspect"], r["zone"]): r for r in saved_schedules_res} if saved_schedules_res else {}
            table_html = """<table style='width:100%; border-collapse: collapse; font-family: sans-serif; text-align: left;'><thead><tr style='background-color: #1f2937; color: #f9fafb;'><th style='padding: 12px; border: 1px solid #374151;'>Aspect Layer Profile</th><th style='padding: 12px; border: 1px solid #374151;'>Target Sector Zone</th><th style='padding: 12px; border: 1px solid #374151;'>Commencement Date</th><th style='padding: 12px; border: 1px solid #374151;'>Wrap Deadline Date</th><th style='padding: 12px; border: 1px solid #374151;'>Target Rate Quantity</th></tr></thead><tbody>"""
            for asp in ["pegging", "piling", "mounting", "modules", "inverter_structure", "inverter", "transformer", "dc_cabling", "ac_cabling"]:
                for zone in clean_zones:
                    match = sched_lookup.get((asp, zone))
                    if match:
                        s_dt = datetime.strptime(match["start_date"], "%Y-%m-%d").date()
                        e_dt = datetime.strptime(match["end_date"], "%Y-%m-%d").date()
                        row_style = "style='background-color: rgba(234, 179, 8, 0.15); font-weight: bold; border-left: 5px solid #eab308;'" if (s_dt <= current_system_date <= e_dt) else ""
                        table_html += f"<tr {row_style}><td style='padding:12px; border:1px solid #374151;'>{asp.upper()}</td><td style='padding:12px; border:1px solid #374151;'>{zone}</td><td style='padding:12px; border:1px solid #374151;'>{match['start_date']}</td><td style='padding:12px; border:1px solid #374151;'>{match['end_date']}</td><td style='padding:12px; border:1px solid #374151;'>{match['daily_target']} Units/Day</td></tr>"
                    else:
                        table_html += f"<tr style='color: #64748b; font-style: italic;'><td style='padding:12px; border:1px solid #374151;'>{asp.upper()}</td><td style='padding:12px; border:1px solid #374151;'>{zone}</td><td style='padding:12px; border:1px solid #374151;'>⏳ Pending</td><td style='padding:12px; border:1px solid #374151;'>⏳ Pending</td><td style='padding:12px; border:1px solid #374151;'>⏱️ Unscheduled</td></tr>"
            table_html += "</tbody></table>"
            st.markdown(table_html, unsafe_allow_html=True)

        # --- CREW TAB 2: ACTIVE TRACKER MATRIX WORKSPACE ---
        with crew_tabs[1]:
            st.markdown("### 🛰️ Live Assembly Deployment Workspace Trackers Console Deck")
            selected_crew_aspect = st.selectbox("Select Active Work Aspect Layer Profile Context:", ["pegging", "piling", "mounting", "modules", "inverter_structure", "inverter", "transformer", "dc_cabling", "ac_cabling"], key="crew_run_layer")
            selected_crew_zone = st.selectbox("Select Active Operational Sector Zone Working Context Area:", clean_zones, key="crew_run_zone")
            
            schedule_meta = supabase.table("project_schedules").select("*").eq("farm_id", st.session_state.active_site_id).eq("aspect", selected_crew_aspect).eq("zone", selected_crew_zone).execute().data
            
            if not schedule_meta:
                st.warning(f"📅 **Awaiting Calendar Broadcast:** Operational timelines for **{selected_crew_aspect.upper()}** inside **{selected_crew_zone}** have not been initialized yet.")
            else:
                sched_bound = schedule_meta[0]
                start_bound_dt = datetime.strptime(sched_bound["start_date"], "%Y-%m-%d").date()
                end_bound_dt = datetime.strptime(sched_bound["end_date"], "%Y-%m-%d").date()
                is_editable_window = (start_bound_dt <= current_system_date <= end_bound_dt)
                
                if not is_editable_window:
                    st.error(f"🔒 **Operational Window Locked:** Editing access is frozen because your current system date ({current_date_str}) falls outside active bounds.")
                
                raw_logs = supabase.table("daily_progress_logs").select("*").eq("farm_id", st.session_state.active_site_id).eq("aspect", selected_crew_aspect).eq("zone", selected_crew_zone).execute().data
                logs_lookup = {r["log_date"]: r for r in raw_logs} if raw_logs else {}
                existing_remark = logs_lookup.get(current_date_str, {}).get("remark", "")
                
                target_runrate = float(sched_bound.get("daily_target", 0.0))

                # Safe topology fetching
                raw_bg = current_farm_record.get("background_image_url") if current_farm_record else None
                try: topo_meta_obj = json.loads(raw_bg) if raw_bg else {}
                except Exception: topo_meta_obj = {}
                
                crew_zone_filtered_blocks = [b for b in active_table_data if str(b.get("assigned_zone")) == str(selected_crew_zone)]
                crew_zone_element_count = 0
                if selected_crew_aspect in ["pegging", "piling"]:
                    for b in crew_zone_filtered_blocks:
                        ev = b.get("section_group") if b.get("section_group") is not None else 403
                        crew_zone_element_count += int((ev // 100) * (ev % 100)) if ev > 100 else 12
                elif selected_crew_aspect in ["mounting", "dc_cabling"]:
                    crew_zone_element_count = len(crew_zone_filtered_blocks)
                elif selected_crew_aspect == "modules":
                    crew_zone_element_count = sum([int((b["max_r"] - b["min_r"] + 1) * (b["max_c"] - b["min_c"] + 1)) for b in crew_zone_filtered_blocks])
                elif selected_crew_aspect in ["inverter_structure", "inverter", "ac_cabling"]:
                    for inv in topo_meta_obj.get("inverters", []):
                        if any(str(b.get("assigned_zone")) == str(selected_crew_zone) and (b["min_c"]*14) <= inv.get("x",0) <= ((b["max_c"]+1)*14) and (b["min_r"]*14) <= inv.get("y",0) <= ((b["max_r"]+1)*14) for b in active_table_data):
                            crew_zone_element_count += 1
                elif selected_crew_aspect == "transformer":
                    crew_zone_element_count = len(topo_meta_obj.get("transformers", []))

                # Canvas UI (stripped of internal HTML inputs/buttons)
                html_crew_engine = """
                <div style="background:#090d16; padding:12px; border-radius:12px; font-family:sans-serif; position:relative; touch-action:none; user-select:none;">
                    <div style="color: #94a3b8; font-size: 13px; margin-bottom: 8px;">
                        🎮 <b>Real-Time Tracking:</b> <span style="color:#22c55e; font-weight:bold;">Left-Click + Drag Box</span> to select items | <span style="color:#ef4444; font-weight:bold;">Single Left-Click</span> to toggle individual nodes. Selections sync to the database instantly.
                    </div>
                    <div style="width:100%; max-height:480px; border:2px solid #1e293b; border-radius:8px; overflow:hidden;">
                        <canvas id="crew_canvas_tracker_element" width="1500" height="480" style="background:#020617; display:block;"></canvas>
                    </div>
                </div>
                <script>
                    (function() {
                        const dataset = JSON.parse(atob("__JSON_DATA_B64__"));
                        const aspect = "ACTIVE_ASPECT_VAL"; const targetZone = "ACTIVE_ZONE_VAL"; const sysDateStr = "SYSTEM_DATE_VAL"; const isEditable = __IS_EDITABLE_VAL__;
                        const canvas = document.getElementById("crew_canvas_tracker_element"); const ctx = canvas.getContext('2d'); const CELL = 14;
                        let minX = MIN_C_VAL, maxX = MAX_C_VAL, minY = MIN_R_VAL, maxY = MAX_R_VAL;
                        const mapWidth = (maxX - minX + 1) * CELL; const mapHeight = (maxY - minY + 1) * CELL;
                        let scale = Math.min((canvas.width - 60) / mapWidth, (canvas.height - 60) / mapHeight);
                        let offsetX = (canvas.width / 2) - (mapWidth * scale / 2) - (minX * CELL * scale);
                        let offsetY = (canvas.height / 2) - (mapHeight * scale / 2) - (minY * CELL * scale);
                        let isPanning = false, isSelecting = false; let sX = 0, sY = 0, cX = 0, cY = 0;
                        
                        canvas.addEventListener('contextmenu', e => e.preventDefault());
                        
                        function getNodeColor(b, aspectKey) {
                            if (b.assigned_zone !== targetZone) return '#222d3d'; 
                            let status = b[aspectKey + '_status'] || 'pending'; let dateVal = b[aspectKey + '_date'];
                            if (status === 'completed') { if (dateVal === sysDateStr) return '#eab308'; return '#22c55e'; }
                            return '#1f2937';
                        }
                        function draw() {
                            ctx.fillStyle = '#020617'; ctx.fillRect(0, 0, canvas.width, canvas.height);
                            ctx.save(); ctx.translate(offsetX, offsetY); ctx.scale(scale, scale);
                            dataset.forEach(b => {
                                let x = b.min_c * CELL; let y = b.min_r * CELL; let w = (b.max_c - b.min_c + 1) * CELL; let h = (b.max_r - b.min_r + 1) * CELL;
                                ctx.fillStyle = getNodeColor(b, aspect); ctx.fillRect(x, y, w, h);
                                ctx.strokeStyle = '#090d16'; ctx.lineWidth = 0.5; ctx.strokeRect(x, y, w, h);
                            });
                            ctx.restore();
                            if (isSelecting && isEditable) { ctx.strokeStyle = '#22c55e'; ctx.lineWidth = 1.5; ctx.strokeRect(sX, sY, cX - sX, cY - sY); }
                        }
                        canvas.addEventListener('mousemove', e => {
                            const rect = canvas.getBoundingClientRect(); const mX = e.clientX - rect.left; const mY = e.clientY - rect.top;
                            if (isPanning) { offsetX = e.clientX - sX; offsetY = e.clientY - sY; draw(); return; }
                            if (isSelecting) { cX = mX; cY = mY; draw(); return; }
                        });
                        canvas.addEventListener('mousedown', e => {
                            const rect = canvas.getBoundingClientRect(); const mX = e.clientX - rect.left; const mY = e.clientY - rect.top;
                            if (e.button === 2) { isPanning = true; sX = e.clientX - offsetX; sY = e.clientY - offsetY; }
                            else if (e.button === 0 && isEditable) { isSelecting = true; sX = mX; sY = mY; cX = mX; cY = mY; }
                        });
                        canvas.addEventListener('mouseup', async (e) => {
                            if (isPanning) { isPanning = false; return; }
                            if (isSelecting) {
                                isSelecting = false; const rect = canvas.getBoundingClientRect(); const mX = e.clientX - rect.left; const mY = e.clientY - rect.top;
                                let wX1 = Math.min((sX - offsetX)/scale, (mX - offsetX)/scale); let wX2 = Math.max((sX - offsetX)/scale, (mX - offsetX)/scale);
                                let wY1 = Math.min((sY - offsetY)/scale, (mY - offsetY)/scale); let wY2 = Math.max((sY - offsetY)/scale, (mY - offsetY)/scale);
                                let isLasso = Math.abs(mX - sX) > 4 || Math.abs(mY - sY) > 4;
                                
                                for (let b of dataset) {
                                    if (b.assigned_zone !== targetZone) continue;
                                    let cx = b.min_c * CELL; let cy = b.min_r * CELL;
                                    let hit = isLasso ? (cx >= wX1 && cx <= wX2 && cy >= wY1 && cy <= wY2) : (wX1 >= cx && wX1 <= (b.max_c+1)*CELL && wY1 >= cy && wY1 <= (b.max_r+1)*CELL);
                                    if (hit) {
                                        let currentStatus = b[aspect + '_status'] || 'pending'; let currentDateVal = b[aspect + '_date'];
                                        if (currentStatus === 'completed' && currentDateVal !== sysDateStr) continue;
                                        
                                        let nextStatus = (currentStatus === 'completed') ? 'pending' : 'completed';
                                        let nextDate = (nextStatus === 'completed') ? sysDateStr : null;
                                        
                                        b[aspect + '_status'] = nextStatus; b[aspect + '_date'] = nextDate;
                                        
                                        // Push updates to the database instantly
                                        fetch("SUPABASE_URL_VAL/rest/v1/structures?id=eq." + b.id, {
                                            method: "PATCH", headers: { "apikey": "SUPABASE_KEY_VAL", "Authorization": "Bearer SUPABASE_KEY_VAL", "Content-Type": "application/json" },
                                            body: JSON.stringify({ [aspect + '_status']: nextStatus, [aspect + '_date']: nextDate })
                                        });
                                    }
                                }
                                draw();
                            }
                        });
                        canvas.addEventListener('wheel', e => {
                            e.preventDefault(); const rect = canvas.getBoundingClientRect(); const mouseX = e.clientX - rect.left; const mouseY = e.clientY - rect.top;
                            const gridX = (mouseX - offsetX) / scale; const gridY = (mouseY - offsetY) / scale;
                            scale *= (e.deltaY < 0 ? 1.15 : 0.85); scale = Math.max(0.01, Math.min(scale, 20));
                            offsetX = mouseX - gridX * scale; offsetY = mouseY - gridY * scale; draw();
                        }, { passive: false });
                        draw();
                    })();
                </script>
                """
                
                # SAFE REPLACEMENT CHAIN (Bypasses NameError completely)
                safe_b64_topo = base64.b64encode((current_farm_record.get("background_image_url") or "{}").encode("utf-8")).decode("utf-8")
                
                html_crew_engine = html_crew_engine.replace("__JSON_DATA_B64__", b64_json_data)
                html_crew_engine = html_crew_engine.replace("__TOPOLOGY_METADATA_B64__", safe_b64_topo)
                html_crew_engine = html_crew_engine.replace("ACTIVE_ASPECT_VAL", selected_crew_aspect)
                html_crew_engine = html_crew_engine.replace("ACTIVE_ZONE_VAL", selected_crew_zone)
                html_crew_engine = html_crew_engine.replace("SYSTEM_DATE_VAL", current_date_str)
                html_crew_engine = html_crew_engine.replace("MIN_C_VAL", str(min_c))
                html_crew_engine = html_crew_engine.replace("MAX_C_VAL", str(max_c))
                html_crew_engine = html_crew_engine.replace("MIN_R_VAL", str(min_r))
                html_crew_engine = html_crew_engine.replace("MAX_R_VAL", str(max_r))
                html_crew_engine = html_crew_engine.replace("SUPABASE_URL_VAL", SUPABASE_URL)
                html_crew_engine = html_crew_engine.replace("SUPABASE_KEY_VAL", SUPABASE_KEY)
                html_crew_engine = html_crew_engine.replace("__IS_EDITABLE_VAL__", "true" if is_editable_window else "false")

                components.html(html_crew_engine, height=520)

                # NATIVE STREAMLIT REMARK & SAVE CONTROLS
                st.write("")
                typed_remark = st.text_input("📝 Append Shift Remarks & Blockage Mitigation Notes:", value=existing_remark, key="crew_live_remark_field_input")
                
                if st.button("💾 Save Target Progress & Shift Logs", type="primary", use_container_width=True, disabled=not is_editable_window):
                    with st.spinner("Synchronizing operations ledger logs..."):
                        # Fetch the live total tally from the database immediately
                        fresh_db_blocks = supabase.table("structures").select("id").eq("farm_id", st.session_state.active_site_id).eq("assigned_zone", selected_crew_zone).eq(f"{selected_crew_aspect}_status", "completed").eq(f"{selected_crew_aspect}_date", current_date_str).execute().data
                        absolute_total_completions = len(fresh_db_blocks) if fresh_db_blocks else 0
                        
                        current_day_target_quota = int(target_runrate)
                        if current_date_str in logs_lookup and logs_lookup[current_date_str].get("target_units", 0) > 0:
                            current_day_target_quota = logs_lookup[current_date_str]["target_units"]
                            
                        computed_deviation = absolute_total_completions - current_day_target_quota
                        
                        # 🎯 FIXED: Proper upsert structure using clean variables and valid header payload modifiers
                        supabase.table("daily_progress_logs").upsert({
                            "farm_id": str(st.session_state.active_site_id), 
                            "aspect": str(selected_crew_aspect), 
                            "zone": str(selected_crew_zone),
                            "log_date": current_date_str, 
                            "target_units": int(current_day_target_quota), 
                            "installed_units": int(absolute_total_completions),
                            "deviation": int(computed_deviation), 
                            "remark": str(typed_remark)
                        }, on_conflict="farm_id, aspect, zone, log_date").execute()
                        
                        st.cache_resource.clear()
                        st.success("🎉 Shift log metrics updated successfully!")
                        time.sleep(0.5)
                        st.rerun()

                st.markdown("#### 📊 Operational Run-Rate Performance Metrics Analytics Calendar")
                table_html_tab2 = """<table style='width:100%; border-collapse: collapse; font-family: sans-serif; text-align: left;'><thead><tr style='background-color: #1f2937; color: #f9fafb;'><th style='padding: 12px; border: 1px solid #374151;'>Date Window</th><th style='padding: 12px; border: 1px solid #374151;'>Production Target</th><th style='padding: 12px; border: 1px solid #374151;'>Assembled Quantity</th><th style='padding: 12px; border: 1px solid #374151;'>Performance Deviation</th><th style='padding: 12px; border: 1px solid #374151;'>Field Remark Notes</th></tr></thead><tbody>"""
                
                total_target_quota = 0.0; total_installed_quota = 0.0; total_deviation_quota = 0.0
                loop_date = start_bound_dt
                while loop_date <= end_bound_dt:
                    loop_date_str = str(loop_date)
                    matched_log = logs_lookup.get(loop_date_str)
                    if loop_date.weekday() < 5 or matched_log:
                        # Pull live completions count from the cached data map
                        inst_count = sum(1 for b in active_table_data if b.get("assigned_zone") == selected_crew_zone and b.get(f"{selected_crew_aspect}_status") == "completed" and b.get(f"{selected_crew_aspect}_date") == loop_date_str)
                        
                        if loop_date_str == current_date_str:
                            remark_display = matched_log.get("remark") if (matched_log and matched_log.get("remark")) else "Real-time field session tracking active."
                            row_style_attr = "style='background-color: rgba(234, 179, 8, 0.22) !important; font-weight: bold !important; border-left: 5px solid #eab308;'"
                        else:
                            remark_display = matched_log.get("remark") if (matched_log and matched_log.get("remark")) else "No operational notes submitted."
                            row_style_attr = ""
                        
                        current_day_target = float(matched_log["target_units"]) if (matched_log and matched_log.get("target_units", 0) > 0) else target_runrate
                        cur_dev = float(inst_count - current_day_target)
                        
                        total_target_quota += current_day_target; total_installed_quota += float(inst_count); total_deviation_quota += cur_dev
                        table_html_tab2 += f"""<tr {row_style_attr}><td style='padding:10px; border:1px solid #374151;'>{loop_date_str}</td><td style='padding:10px; border:1px solid #374151;'>{round(current_day_target)} Units</td><td style='padding:10px; border:1px solid #374151;'>{int(inst_count)} Units</td><td style='padding:10px; border:1px solid #374151;'>{"🟢 +" if cur_dev >= 0 else "🔴 "}{round(cur_dev)}</td><td style='padding:10px; border:1px solid #374151;'>{remark_display}</td></tr>"""
                    loop_date += timedelta(days=1)
                
                if abs(total_target_quota - crew_zone_element_count) <= 5:
                    total_target_quota = crew_zone_element_count
                    total_deviation_quota = total_installed_quota - total_target_quota

                table_html_tab2 += f"""<tr style='background:#111827; font-weight:bold;'><td style='padding:12px; border:1px solid #374151;'>📊 RUNRATES FOOTPRINT ROLLUP TOTALS</td><td style='padding:12px; border:1px solid #374151;'>{round(total_target_quota)} Units</td><td style='padding:12px; border:1px solid #374151;'>{round(total_installed_quota)} Units</td><td style='padding:12px; border:1px solid #374151;'>{"🟢 +" if total_deviation_quota >= 0 else "🔴 "}{round(total_deviation_quota)}</td><td style='padding:12px; border:1px solid #374151;'>🏁 Master Balance Ledger Log</td></tr></tbody></table>"""
                st.markdown(table_html_tab2, unsafe_allow_html=True)

        # --- CREW TAB 3: FIELD SHIFT HISTORY LOG VIEWER ---
        with crew_tabs[2]:
            st.markdown("### 🕒 Historic Operational Field Execution Ledger Lookups")
            st.caption("Select an operational layer aspect and historical date point below to evaluate precisely what milestones your crew captured during that shift.")
            
            hist_cols = st.columns(2)
            with hist_cols[0]: lookup_crew_aspect = st.selectbox("Choose Historical Aspect Layer:", ["pegging", "piling", "mounting", "modules", "inverter_structure", "inverter", "transformer", "dc_cabling", "ac_cabling"], key="crew_hist_lookup_asp")
            with hist_cols[1]: lookup_crew_date = st.date_input("Target Evaluation Shift Date Window:", value=current_system_date, key="crew_hist_lookup_dt")
            
            lookup_date_str = str(lookup_crew_date)

            html_hist_zone_map = """
            <div style="background:#090d16; padding:12px; border-radius:12px; position:relative; touch-action:none; user-select: none; font-family:sans-serif;">
                <div style="color: #94a3b8; font-size: 13px; margin-bottom: 8px;">
                    🗺️ <b>Playback Controls:</b> Move cursor to check unit details | <span style="color:#22c55e; font-weight:bold;">Green Cells</span> represent past completions | <span style="color:#eab308; font-weight:bold;">Yellow Cells</span> indicate items installed on this specific shift window. Right-Click + Drag to Pan | Scroll to Zoom.
                </div>
                <div id="crew_hist_canvas_tooltip" style="position: absolute; display: none; background: rgba(15, 23, 42, 0.95); color: #f8fafc; border: 1px solid #3b82f6; padding: 6px 12px; border-radius: 4px; font-size: 12px; pointer-events: none; z-index: 99999; box-shadow: 0 4px 12px rgba(0,0,0,0.5); font-weight: bold;"></div>
                <div style="width:100%; max-height:350px; border:2px solid #1e293b; border-radius:8px; overflow:hidden;">
                    <canvas id="crew_hist_overview_canvas" width="1500" height="350" style="background:#020617; display:block;"></canvas>
                </div>
            </div>
            <script>
                (function() {
                    const blocks = JSON.parse(atob("__JSON_DATA_B64__"));
                    const canvas = document.getElementById("crew_hist_overview_canvas"); const ctx = canvas.getContext('2d'); const CELL = 14;
                    const tooltip = document.getElementById("crew_hist_canvas_tooltip");
                    const aspect = "ACTIVE_ASPECT_VAL"; const evaluationDateStr = "EVALUATION_DATE_VAL";
                    
                    let minX = MIN_C_VAL, maxX = MAX_C_VAL, minY = MIN_R_VAL, maxY = MAX_R_VAL;
                    const mapWidth = (maxX - minX + 1) * CELL; const mapHeight = (maxY - minY + 1) * CELL;
                    let scale = Math.min((canvas.width - 60) / mapWidth, (canvas.height - 60) / mapHeight);
                    let offsetX = (canvas.width / 2) - (mapWidth * scale / 2) - (minX * CELL * scale);
                    let offsetY = (canvas.height / 2) - (mapHeight * scale / 2) - (minY * CELL * scale);
                    let isPanning = false; let startX = 0, startY = 0;
                    
                    function draw() {
                        ctx.clearRect(0, 0, canvas.width, canvas.height); ctx.save(); ctx.translate(offsetX, offsetY); ctx.scale(scale,scale);
                        blocks.forEach(b => {
                            let status = b[aspect + '_status'] || 'pending'; let compDate = b[aspect + '_date'];
                            if (status === 'completed' && compDate) {
                                if (compDate === evaluationDateStr) ctx.fillStyle = '#eab308';
                                else if (compDate < evaluationDateStr) ctx.fillStyle = '#22c55e';
                                else ctx.fillStyle = '#1e293b';
                            } else ctx.fillStyle = '#1e293b';
                            
                            let x = b.min_c * CELL; let y = b.min_r * CELL; let w = (b.max_c - b.min_c + 1) * CELL; let h = (b.max_r - b.min_r + 1) * CELL;
                            ctx.fillRect(x, y, w, h); ctx.strokeStyle = '#020617'; ctx.lineWidth = 0.5; ctx.strokeRect(x, y, w, h);
                        });
                        ctx.restore();
                    }
                    canvas.addEventListener('mousemove', e => {
                        const rect = canvas.getBoundingClientRect(); const mX = e.clientX - rect.left; const mY = e.clientY - rect.top;
                        if (isPanning) { offsetX = e.clientX - startX; offsetY = e.clientY - startY; draw(); tooltip.style.display = "none"; return; }
                        let worldX = (mX - offsetX) / scale; let worldY = (mY - offsetY) / scale;
                        let hoveredBlock = null;
                        for (let b of blocks) {
                            let x = b.min_c * CELL; let y = b.min_r * CELL; let w = (b.max_c - b.min_c + 1) * CELL; let h = (b.max_r - b.min_r + 1) * CELL;
                            if (worldX >= x && worldX <= x + w && worldY >= y && worldY <= y + h) { hoveredBlock = b; break; }
                        }
                        if (hoveredBlock) {
                            tooltip.style.display = "block"; tooltip.style.left = (mX + 15) + "px"; tooltip.style.top = (mY + 15) + "px";
                            let status = hoveredBlock[aspect + '_status'] || 'pending'; let compDate = hoveredBlock[aspect + '_date'] || 'N/A';
                            tooltip.innerHTML = `📍 Unit: ${hoveredBlock.table_label}<br>🗺️ Zone: ${hoveredBlock.assigned_zone}<br>⚡ Status: ${status.toUpperCase()}<br>📅 Built Date: ${compDate}`;
                        } else tooltip.style.display = "none";
                    });
                    canvas.addEventListener('mousedown', e => { if (e.button === 2) { isPanning = true; startX = e.clientX - offsetX; startY = e.clientY - offsetY; tooltip.style.display = "none"; } });
                    canvas.addEventListener('mouseup', () => { isPanning = false; });
                    canvas.addEventListener('wheel', e => {
                        e.preventDefault(); const rect = canvas.getBoundingClientRect(); const mouseX = e.clientX - rect.left; const mouseY = e.clientY - rect.top;
                        const gridX = (mouseX - offsetX) / scale; const gridY = (mouseY - offsetY) / scale;
                        scale *= (e.deltaY < 0 ? 1.15 : 0.85); scale = Math.max(0.01, Math.min(scale, 20));
                        offsetX = mouseX - gridX * scale; offsetY = mouseY - gridY * scale; draw();
                    }, { passive: false });
                    draw();
                })();
            </script>
            """
            html_hist_zone_map = html_hist_zone_map.replace("__JSON_DATA_B64__", b64_json_data).replace("ACTIVE_ASPECT_VAL", lookup_crew_aspect).replace("EVALUATION_DATE_VAL", lookup_date_str).replace("MIN_C_VAL", str(min_c)).replace("MAX_C_VAL", str(max_c)).replace("MIN_R_VAL", str(min_r)).replace("MAX_R_VAL", str(max_r))
            components.html(html_hist_zone_map, height=390)

            st.markdown("#### 📅 Comprehensive Historical Performance Calendars")
            hist_sched_records = supabase.table("project_schedules").select("*").eq("farm_id", st.session_state.active_site_id).eq("aspect", lookup_crew_aspect).execute().data
            
            if hist_sched_records:
                raw_hist_logs = supabase.table("daily_progress_logs").select("*").eq("farm_id", st.session_state.active_site_id).eq("aspect", lookup_crew_aspect).execute().data
                hist_logs_lookup = {(r["zone"], r["log_date"]): r for r in raw_hist_logs} if raw_hist_logs else {}
                
                try: topo_meta_obj = json.loads(current_farm_record.get("background_image_url") or "{}")
                except Exception: topo_meta_obj = {}

                # Stack multiple historical tracker sheets back-to-back dynamically on one page layout as requested
                for zone_name in clean_zones:
                    sched = next((s for s in hist_sched_records if s["zone"] == zone_name), None)
                    if not sched: continue
                    
                    s_bound = datetime.strptime(sched["start_date"], "%Y-%m-%d").date()
                    e_bound = datetime.strptime(sched["end_date"], "%Y-%m-%d").date()
                    t_runrate = float(sched.get("daily_target", 0.0))
                    
                    st.markdown(f"##### 🗺️ Performance Metrics Log Schedule: **{zone_name.upper()}**")
                    
                    loop_filtered_blocks = [b for b in active_table_data if str(b.get("assigned_zone")) == str(zone_name)]
                    z_element_count = 0
                    if lookup_crew_aspect in ["pegging", "piling"]:
                        for b in loop_filtered_blocks:
                            ev = b.get("section_group") if b.get("section_group") is not None else 403
                            z_element_count += int((ev // 100) * (ev % 100)) if ev > 100 else 12
                    elif lookup_crew_aspect in ["mounting", "dc_cabling"]:
                        z_element_count = len(loop_filtered_blocks)
                    elif lookup_crew_aspect == "modules":
                        z_element_count = sum([int((b["max_r"] - b["min_r"] + 1) * (b["max_c"] - b["min_c"] + 1)) for b in loop_filtered_blocks])
                    elif lookup_crew_aspect in ["inverter_structure", "inverter", "ac_cabling"]:
                        for inv in topo_meta_obj.get("inverters", []):
                            if any(str(b.get("assigned_zone")) == str(zone_name) and (b["min_c"]*14) <= inv.get("x",0) <= ((b["max_c"]+1)*14) and (b["min_r"]*14) <= inv.get("y",0) <= ((b["max_r"]+1)*14) for b in active_table_data):
                                z_element_count += 1
                    elif lookup_crew_aspect == "transformer":
                        z_element_count = len(topo_meta_obj.get("transformers", []))

                    hist_table_html = """<table style='width:100%; border-collapse: collapse; font-family: sans-serif; text-align: left; margin-bottom: 24px;'>
                        <thead>
                            <tr style='background-color: #1f2937; color: #f9fafb;'>
                                <th style='padding: 12px; border: 1px solid #374151;'>Date Window</th>
                                <th style='padding: 12px; border: 1px solid #374151;'>Production Target</th>
                                <th style='padding: 12px; border: 1px solid #374151;'>Actual Installed</th>
                                <th style='padding: 12px; border: 1px solid #374151;'>Shift Performance Deviation</th>
                                <th style='padding: 12px; border: 1px solid #374151;'>Supervisor Operational Remarks</th>
                            </tr>
                        </thead><tbody>"""

                    z_total_target = 0.0; z_total_installed = 0.0; z_total_deviation = 0.0
                    loop_dt = s_bound
                    while loop_dt <= e_bound:
                        loop_dt_str = str(loop_dt)
                        matched_log = hist_logs_lookup.get((zone_name, loop_dt_str))
                        
                        if loop_dt.weekday() < 5 or matched_log:
                            inst_val = sum(1 for b in active_table_data if b.get("assigned_zone") == zone_name and b.get(f"{lookup_crew_aspect}_status") == "completed" and b.get(f"{lookup_crew_aspect}_date") == loop_dt_str)
                            t_val = float(matched_log["target_units"]) if (matched_log and matched_log.get("target_units", 0) > 0) else t_runrate
                            dev_val = float(inst_val - t_val)
                            rem_val = matched_log.get("remark") or "No field remarks submitted." if matched_log else "No logs recorded."
                            
                            if loop_dt_str == lookup_date_str:
                                row_style = "style='background-color: rgba(59, 130, 246, 0.24) !important; font-weight: bold !important; border-left: 5px solid #3b82f6 !important;'"
                            else:
                                row_style = ""
                                
                            z_total_target += t_val
                            z_total_installed += float(inst_val)
                            z_total_deviation += dev_val
                            
                            hist_table_html += f"""<tr {row_style}>
                                <td style='padding: 10px; border: 1px solid #374151;'>{loop_dt_str}</td>
                                <td style='padding: 10px; border: 1px solid #374151;'>{round(t_val)} Units</td>
                                <td style='padding: 10px; border: 1px solid #374151;'>{int(inst_val)} Units</td>
                                <td style='padding: 10px; border: 1px solid #374151;'>{"🟢 +" if dev_val >= 0 else "🔴 "}{round(dev_val)}</td>
                                <td style='padding: 10px; border: 1px solid #374151;'>{rem_val}</td>
                            </tr>"""
                        loop_dt += timedelta(days=1)
                        
                    if abs(z_total_target - z_element_count) <= 5:
                        z_total_target = z_element_count
                        z_total_deviation = z_total_installed - z_total_target

                    hist_table_html += f"""<tr style='background:#111827; font-weight:bold;'>
                        <td style='padding:12px; border: 1px solid #374151;'>📊 {zone_name.upper()} SUMMATION ROLLUP TOTALS</td>
                        <td style='padding:12px; border: 1px solid #374151;'>{round(z_total_target)} Units</td>
                        <td style='padding:12px; border: 1px solid #374151;'>{round(z_total_installed)} Units</td>
                        <td style='padding:12px; border: 1px solid #374151;'>{"🟢 +" if z_total_deviation >= 0 else "🔴 "}{round(z_total_deviation)}</td>
                        <td style='padding:12px; border: 1px solid #374151;'>🏁 Zone Specific Ledger Balance</td>
                    </tr></tbody></table>"""
                    st.markdown(hist_table_html, unsafe_allow_html=True)
            else:
                st.info("ℹ️ No active milestone configuration layers match this aspect layer context.")
